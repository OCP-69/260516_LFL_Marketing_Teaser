#!/usr/bin/env python3
"""Build Email_Outreach_Interpack_Warm worksheet in CRM funnel xlsx."""
import csv
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl required: pip install openpyxl")

BASE = Path(__file__).resolve().parent
CSV_PATH = BASE / "260609_Email_Kontakte_Interpack_Extended_v2.csv"
XLSX_PATH = BASE / "260605_CRM_Funnel_Interpack_List.xlsx"
SHEET_NAME = "Email_Outreach_Interpack_Warm"
TEASER_PATH = (
    "02_Aktuell/Teaser/Variante_E_Interpack/"
    "LFL_Teaser_Variante_E_Interpack_A4_v02.pptx"
)

KONZERN = {1, 5, 9, 10, 11, 13, 15, 18}
ICP_MITTELSTAND = {2, 3, 4, 6, 7, 8, 12, 14, 19}

COMPANY_META = {
    1: {
        "short": "Beumer Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Mehr Angebotsvarianten bei gleicher Engineering-Kapazität — Forge Engine rechnet BOM- und CAD-Änderungen im Design in belastbare Kosten, ohne ERP-Rollout.",
        "benefit_eng_de": "Weniger Iterationsschleifen vor Freigabe: Varianten aus BOM, CAD und Einkauf sofort in Euro — kein neues ERP.",
        "benefit_ceo_en": "More quote variants with the same engineering capacity — Forge Engine turns BOM and CAD changes into reliable costs at design stage, without an ERP rollout.",
        "benefit_eng_en": "Fewer iteration loops before release: BOM, CAD and purchasing scenarios in euros instantly — no new ERP.",
        "prio": "C",
    },
    2: {
        "short": "Handtmann Processing",
        "mouaz_de": "Kollegen von Ihrem Messestand",
        "mouaz_en": "colleagues at your stand",
        "benefit_ceo_de": "Kostensicherheit vor Freigabe bei Varianten — typisch weniger teure Nacharbeit und klarere Margen pro Auftrag.",
        "benefit_eng_de": "Material- und Prozess-What-ifs direkt in der Konstruktion in Euro — weniger Rückfragen an Einkauf und Kalkulation.",
        "benefit_ceo_en": "Cost certainty before release on variants — typically less costly rework and clearer margins per order.",
        "benefit_eng_en": "Material and process what-ifs in euros during design — fewer back-and-forths with purchasing and costing.",
        "prio": "A",
        "visitenkarte": "ja (Stand-Gespräch)",
    },
    3: {
        "short": "Winkler+Dünnebier",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Schnellere RFQs bei Sondermaschinen — Varianten im Design kalkulieren statt Wochen in Excel-Schleifen.",
        "benefit_eng_de": "Stücklisten- und CAD-Varianten vor Freigabe durchrechnen — weniger Überraschungen in der Fertigung.",
        "benefit_ceo_en": "Faster RFQs on custom machinery — cost variants in design instead of weeks in spreadsheet loops.",
        "benefit_eng_en": "Run BOM and CAD variants before release — fewer surprises in production.",
        "prio": "A",
    },
    4: {
        "short": "KÖRA-PACKMAT",
        "mouaz_de": "Herrn Faller (Einkauf & Materialwirtschaft)",
        "mouaz_en": "Mr Faller (purchasing & materials)",
        "benefit_ceo_de": "Lieferanten- und Materialszenarien schon im Design sichtbar — bessere Angebotsbasis bei kundenspezifischen Maschinen.",
        "benefit_eng_de": "Einkaufspreise und BOM-Alternativen im Konstruktionsmoment vergleichen — ohne ERP-Projekt.",
        "benefit_ceo_en": "Supplier and material scenarios visible at design stage — stronger quote basis on custom machinery.",
        "benefit_eng_en": "Compare purchase prices and BOM alternatives during design — without an ERP project.",
        "prio": "A",
        "visitenkarte": "ja (Oliver Faller)",
    },
    5: {
        "short": "Syntegon",
        "mouaz_de": "Herrn Ruener von Ihrem Team",
        "mouaz_en": "Mr Ruener from your team",
        "benefit_ceo_de": "Nach der Messe: Variantenflut schneller bewerten — Engineering-Kapazität für Angebote statt für Nachkalkulation.",
        "benefit_eng_de": "Sekundärverpackungs-Varianten im Design in Kosten übersetzen — weniger Freigabe-Runden mit Vertrieb.",
        "benefit_ceo_en": "Post-show: evaluate variant floods faster — engineering capacity for quotes, not rework costing.",
        "benefit_eng_en": "Turn secondary packaging variants into costs at design — fewer release rounds with sales.",
        "prio": "B",
        "visitenkarte": "ja (Stefan Ruener)",
    },
    6: {
        "short": "Schubert Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Modulare Linien schneller kalkulierbar — Varianten aus bestehenden BOM- und CAD-Daten, kein ERP-Ersatz.",
        "benefit_eng_de": "Pick-and-place- und Modul-Varianten vor Freigabe in Euro — typisch Tage statt Wochen pro RFQ-Schleife.",
        "benefit_ceo_en": "Modular lines costed faster — variants from existing BOM and CAD data, not an ERP replacement.",
        "benefit_eng_en": "Pick-and-place and module variants in euros before release — typically days not weeks per RFQ loop.",
        "prio": "A",
    },
    7: {
        "short": "SMI Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Schnellere Angebote auf PET-/Verpackungslinien — Kosten aus BOM und CAD schon im Engineering.",
        "benefit_eng_de": "RFQ-Varianten für Verpackungslinien direkt aus Stückliste und CAD — ohne natives CAD im MVP.",
        "benefit_ceo_en": "Faster quotes on PET/packaging lines — costs from BOM and CAD already in engineering.",
        "benefit_eng_en": "RFQ variants for packaging lines straight from BOM and CAD — no native CAD required in the pilot.",
        "prio": "B",
        "lang": "en",
    },
    8: {
        "short": "Romaco Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Pharma-Projektvarianten früher in belastbare Kosten — weniger Risiko in Angebot und Marge.",
        "benefit_eng_de": "Varianten und Compliance-relevante Änderungen im Design durchrechnen — Pilot startet mit wenigen Stücklisten.",
        "benefit_ceo_en": "Pharma project variants in reliable costs earlier — less risk in quote and margin.",
        "benefit_eng_en": "Cost compliance-relevant design changes at design stage — pilot starts with a few BOMs.",
        "prio": "B",
    },
    9: {
        "short": "MULTIVAC",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Integrierte Verpackungslösungen schneller kalkulieren — Engineering entlasten, RFQ-Tempo erhöhen.",
        "benefit_eng_de": "Linien- und Modulvarianten aus BOM, CAD und Einkauf in einer Kostensicht — kein ERP-Rollout.",
        "benefit_ceo_en": "Cost integrated packaging solutions faster — free up engineering, increase RFQ speed.",
        "benefit_eng_en": "Line and module variants from BOM, CAD and purchasing in one cost view — no ERP rollout.",
        "prio": "C",
    },
    10: {
        "short": "Sidel",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Equipment-Portfolio-Varianten schneller bewertbar — Kosten im Design statt erst nach Freigabe.",
        "benefit_eng_de": "Kunden-Spezifikationen und Portfolio-Varianten im Engineering in Euro — weniger Nacharbeit.",
        "benefit_ceo_en": "Evaluate equipment portfolio variants faster — costs at design stage, not after release.",
        "benefit_eng_en": "Customer specs and portfolio variants in euros during engineering — less rework.",
        "prio": "C",
        "lang": "en",
    },
    11: {
        "short": "Krones",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Abfüll- und Verpackungsprojekte: Varianten früher in belastbare Kosten — mehr Angebotskapazität.",
        "benefit_eng_de": "Linienvarianten aus BOM und CAD vor Freigabe — typisch weniger Iteration mit Vertrieb und Einkauf.",
        "benefit_ceo_en": "Filling and packaging projects: reliable costs on variants earlier — more quote capacity.",
        "benefit_eng_en": "Line variants from BOM and CAD before release — typically fewer iterations with sales and purchasing.",
        "prio": "C",
    },
    12: {
        "short": "KHS",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Linienengineering schneller in Angebote übersetzen — Variantenkosten im Design, nicht erst im Nachgang.",
        "benefit_eng_de": "Cross-Function-Varianten aus Stückliste und CAD sofort kalkulieren — Pilot mit wenigen BOMs.",
        "benefit_ceo_en": "Turn line engineering into quotes faster — variant costs at design, not after the fact.",
        "benefit_eng_en": "Cost cross-function variants from BOM and CAD instantly — pilot with a few BOMs.",
        "prio": "A",
    },
    13: {
        "short": "IMA Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Verpackungsmaschinen-Varianten schneller in Angebote — Digitalisierung ohne ERP-Großprojekt.",
        "benefit_eng_de": "Industry-4.0-Daten (BOM, CAD, Einkauf) zu Kostenszenarien im Design — EU-Datenhaltung, DSGVO.",
        "benefit_ceo_en": "Packaging machine variants into quotes faster — digitalisation without a major ERP project.",
        "benefit_eng_en": "Industry 4.0 data (BOM, CAD, purchasing) as cost scenarios at design — EU data hosting, GDPR.",
        "prio": "C",
        "lang": "en",
    },
    14: {
        "short": "Harro Höfliger",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Sonderanlagen: Kosten vor Freigabe belastbar — weniger Marge-Risiko bei kundenspezifischen Projekten.",
        "benefit_eng_de": "Varianten und Freigabe-Entscheidungen mit Euro-Zahlen aus BOM und CAD untermauern.",
        "benefit_ceo_en": "Special-purpose machines: reliable costs before release — less margin risk on custom projects.",
        "benefit_eng_en": "Back variant and release decisions with euro figures from BOM and CAD.",
        "prio": "A",
    },
    15: {
        "short": "GEA Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Prozesstechnik-Varianten schneller in Angebote — Engineering-Kapazität für Innovation statt Nachkalkulation.",
        "benefit_eng_de": "Angebotsvarianten aus bestehenden Stücklisten und Einkaufsdaten — Intelligenzschicht, kein ERP-Ersatz.",
        "benefit_ceo_en": "Process technology variants into quotes faster — engineering for innovation, not rework costing.",
        "benefit_eng_en": "Quote variants from existing BOMs and purchasing data — intelligence layer, not ERP replacement.",
        "prio": "C",
    },
    16: {
        "short": "Fette Compacting",
        "mouaz_de": "Kollegen von Ihrem Messestand",
        "mouaz_en": "colleagues at your stand",
        "benefit_ceo_de": "Tablettier-Varianten früher in Kosten — auch bei Pharma-Segmenten weniger Überraschung nach Freigabe.",
        "benefit_eng_de": "Maschinen- und Werkzeugvarianten im Design durchrechnen — Pilot Q2/2026, wenige Stücklisten.",
        "benefit_ceo_en": "Tableting variants in costs earlier — fewer surprises after release, even in pharma segments.",
        "benefit_eng_en": "Cost machine and tooling variants at design — Q2/2026 pilot, a few BOMs.",
        "prio": "C",
        "hinweis": "Pharma-Segment; niedrige Priorität",
    },
    17: {
        "short": "PolTech Packaging",
        "prio": "AUSSETZEN",
        "hinweis": "Firma nicht identifiziert — CRM klären",
    },
    18: {
        "short": "TOMRA",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Food-Sortier-Projekte: Variantenkosten früher sichtbar — schnellere Angebotsentscheidungen.",
        "benefit_eng_de": "Equipment-Varianten aus BOM und Engineering-Daten — ohne ERP-Rollout.",
        "benefit_ceo_en": "Food sorting projects: variant costs visible earlier — faster quote decisions.",
        "benefit_eng_en": "Equipment variants from BOM and engineering data — without an ERP rollout.",
        "prio": "C",
        "lang": "en",
    },
    19: {
        "short": "Uhlmann Pac-Systeme",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "benefit_ceo_de": "Pharma-Verpackungslinien: Varianten und Margen vor Freigabe klarer — weniger teure Änderungen im Projekt.",
        "benefit_eng_de": "Linien- und Modulvarianten im Solution Engineering direkt in Euro — DSGVO, Daten beim Kunden.",
        "benefit_ceo_en": "Pharma packaging lines: clearer variants and margins before release — fewer costly project changes.",
        "benefit_eng_en": "Line and module variants in solution engineering in euros — GDPR, data stays on your side.",
        "prio": "A",
    },
}


def is_german_company(nr: int, location: str) -> bool:
    if COMPANY_META.get(nr, {}).get("lang") == "en":
        return False
    if COMPANY_META.get(nr, {}).get("lang") == "de":
        return True
    loc = (location or "").upper()
    if ", DE" in loc or loc.endswith(" DE") or " DE)" in loc:
        return True
    if " IT" in loc or ", IT" in loc or " FR" in loc or ", FR" in loc:
        return False
    if " NO" in loc or "INTL" in loc:
        return False
    return ", DE" in loc


def salutation_de(name: str) -> str:
    if name in ("k.A.", "", None) or not name:
        return "Guten Tag,"
    if "Jessica" in name:
        return f"Guten Tag Frau {name.split()[-1]},"
    if name.startswith("Dr."):
        parts = name.replace("Dr.-Ing.", "Dr.").split()
        return f"Guten Tag Herr Dr. {parts[-1]},"
    return f"Guten Tag Herr {name.split()[-1]},"


def salutation_en(name: str) -> str:
    if name in ("k.A.", "", None) or not name:
        return "Dear Sir or Madam,"
    parts = name.replace("Dr.-Ing.", "Dr.").split()
    last = parts[-1]
    if "Jessica" in name:
        return f"Dear Ms {last},"
    if name.startswith("Dr."):
        return f"Dear Dr {last},"
    return f"Dear Mr {last},"


def pick_ansatz(nr: int, contact_type: str) -> str:
    if nr == 17:
        return "AUSSETZEN"
    meta = COMPANY_META.get(nr, {})
    if meta.get("prio") == "C" or nr in KONZERN:
        return "A"
    if nr in ICP_MITTELSTAND:
        return "B"
    return "A"


def betreff(short: str, contact_type: str, de: bool, ansatz: str) -> str:
    if de:
        if ansatz == "A":
            return (
                f"Nach interpack — Kostensicherheit bei {short}"
                if contact_type == "CEO/GF"
                else f"Nach interpack — Varianten-Kosten bei {short}"
            )
        return (
            f"interpack-Follow-up + Teaser ({short})"
            if contact_type == "CEO/GF"
            else f"Nach interpack — BOM-Kosten im Design ({short})"
        )
    if ansatz == "A":
        return (
            f"After interpack — cost certainty at {short}"
            if contact_type == "CEO/GF"
            else f"After interpack — variant costing at {short}"
        )
    return (
        f"interpack follow-up + teaser ({short})"
        if contact_type == "CEO/GF"
        else f"After interpack — BOM costs at design ({short})"
    )


def signoff(de: bool) -> str:
    if de:
        return (
            "Beste Grüße\n"
            "Olaf Pick\n"
            "LoopForgeLab GbR\n"
            "olaf@loopforgelab.com"
        )
    return (
        "Best regards\n"
        "Olaf Pick\n"
        "LoopForgeLab GbR\n"
        "olaf@loopforgelab.com"
    )


def body_a(row: dict, meta: dict, de: bool) -> str:
    short = meta["short"]
    is_ceo = row["contact_type"] == "CEO/GF"
    sal = salutation_de(row["name"]) if de else salutation_en(row["name"])
    mouaz = meta["mouaz_de"] if de else meta["mouaz_en"]
    benefit = meta["benefit_ceo_de" if is_ceo else "benefit_eng_de"] if de else meta[
        "benefit_ceo_en" if is_ceo else "benefit_eng_en"
    ]
    sig = signoff(de)

    if de:
        ctx = (
            f"Mein Kollege Mouaz Al-Qudsi (CTO) sprach auf der interpack mit {mouaz}."
        )
        if is_ceo:
            ask = (
                f"Limitiert {short} 2026 eher Angebotskapazität oder Marge pro Auftrag? "
                f"Wenn relevant: 20 Min. — https://calendly.com/olaf-pick/olaf-15min"
            )
        else:
            ask = (
                "Gehören teure Änderungen nach Freigabe bei Ihnen zum Alltag — oder eher zur Ausnahme? "
                "20 Min. reichen für einen ersten Abgleich: https://calendly.com/olaf-pick/olaf-15min"
            )
        return f"{sal}\n\n{benefit}\n\n{ctx} Ich melde mich kurz im Anschluss.\n\n{ask}\n\n{sig}"

    ctx = "My colleague Mouaz Al-Qudsi (CTO) spoke with {mouaz} at interpack.".format(
        mouaz=mouaz
    )
    if is_ceo:
        ask = (
            f"What limits {short} more in 2026 — quote capacity or margin per order? "
            f"If relevant: 20 min — https://calendly.com/olaf-pick/olaf-15min"
        )
    else:
        ask = (
            "Are costly post-release changes routine for you — or the exception? "
            "20 minutes is enough for a first alignment: https://calendly.com/olaf-pick/olaf-15min"
        )
    return f"{sal}\n\n{benefit}\n\n{ctx} I'm following up briefly.\n\n{ask}\n\n{sig}"


def body_b(row: dict, meta: dict, de: bool) -> str:
    short = meta["short"]
    is_ceo = row["contact_type"] == "CEO/GF"
    sal = salutation_de(row["name"]) if de else salutation_en(row["name"])
    mouaz = meta["mouaz_de"] if de else meta["mouaz_en"]
    benefit = meta["benefit_ceo_de" if is_ceo else "benefit_eng_de"] if de else meta[
        "benefit_ceo_en" if is_ceo else "benefit_eng_en"
    ]
    vk = meta.get("visitenkarte")
    sig = signoff(de)

    if de:
        vk_line = " (Visitenkarte liegt bei uns)" if vk else ""
        ctx = f"Mouaz war auf der interpack mit {mouaz} im Gespräch{vk_line}."
        if is_ceo:
            ask = (
                f'Ist "Kosten vor Freigabe" bei {short} ein Top-3-Thema 2026? '
                f"Anhang: interpack-Teaser (2 Min.). 20 Min. KW 25/26 oder: "
                f"https://calendly.com/olaf-pick/olaf-15min"
            )
        else:
            ask = (
                "Passt das zu Konstruktion/Freigabe bei Ihnen? Anhang: Teaser (80 % Kosten im Design). "
                "20 Min.: https://calendly.com/olaf-pick/olaf-15min"
            )
        pilot = "Pilot Q2/2026 · kein ERP-Ersatz · DSGVO · EU-Datenhaltung."
        return f"{sal}\n\n{benefit}\n\n{ctx}\n\n{pilot}\n\n{ask}\n\n{sig}"

    vk_line = " (we have your business card)" if vk else ""
    ctx = f"Mouaz spoke with {mouaz} at interpack{vk_line}."
    if is_ceo:
        ask = (
            f"Is “cost before release” a top-3 topic for {short} in 2026? "
            f"Attached: interpack teaser (2 min read). 20 min, week 25/26 or: "
            f"https://calendly.com/olaf-pick/olaf-15min"
        )
    else:
        ask = (
            "Does this fit design/release at your end? Attached: teaser (80% of costs in design). "
            "20 min: https://calendly.com/olaf-pick/olaf-15min"
        )
    pilot = "Q2/2026 pilot · not an ERP replacement · GDPR · EU data hosting."
    return f"{sal}\n\n{benefit}\n\n{ctx}\n\n{pilot}\n\n{ask}\n\n{sig}"


def main():
    rows_out = []
    with CSV_PATH.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            nr = int(row["nr"])
            meta = COMPANY_META.get(
                nr,
                {
                    "short": row["company"],
                    "mouaz_de": "Kollegen am Stand",
                    "mouaz_en": "colleagues at the stand",
                    "benefit_ceo_de": "Varianten im Design in belastbare Kosten — ohne ERP-Rollout.",
                    "benefit_eng_de": "BOM- und CAD-Varianten vor Freigabe in Euro.",
                    "benefit_ceo_en": "Reliable costs on variants at design — without an ERP rollout.",
                    "benefit_eng_en": "BOM and CAD variants in euros before release.",
                    "prio": "B",
                },
            )
            ansatz = pick_ansatz(nr, row["contact_type"])
            de = is_german_company(nr, row.get("location", ""))
            sprache = "DE" if de else "EN"

            if ansatz == "AUSSETZEN":
                rows_out.append({
                    **row,
                    "sprache": "—",
                    "ansatz": ansatz,
                    "mouaz_kontakt": "—",
                    "betreff": "—",
                    "email_body": "— AUSSETZEN: Firma nicht identifiziert — bitte CRM klären",
                    "anhang": "nein",
                    "teaser_pfad": "—",
                    "versand_status": "aussetzen",
                    "icp_prio": "—",
                    "template_ref": "—",
                    "notizen": row.get("notes", ""),
                    "notizen_extra": meta.get("hinweis", ""),
                })
                continue

            if ansatz == "A":
                subj = betreff(meta["short"], row["contact_type"], de, "A")
                body = body_a(row, meta, de)
            else:
                subj = betreff(meta["short"], row["contact_type"], de, "B")
                body = body_b(row, meta, de)
            anhang = "ja" if re.search(
                r"\bAnhang\b|\bAttached\b|angehängt|im Anhang|teaser attached|enclosed",
                body,
                re.I,
            ) else "nein"

            rows_out.append({
                **row,
                "sprache": sprache,
                "ansatz": ansatz,
                "mouaz_kontakt": meta["mouaz_de"] if de else meta["mouaz_en"],
                "betreff": subj,
                "email_body": body,
                "anhang": anhang,
                "teaser_pfad": TEASER_PATH if anhang == "ja" else "",
                "versand_status": "vorbereitet",
                "icp_prio": meta.get("prio", "B"),
                "template_ref": f"Ansatz_{ansatz}_kurz_{sprache}",
                "notizen_extra": meta.get("hinweis", meta.get("visitenkarte", "")),
                "notizen": row.get("notes", ""),
            })

    headers = [
        "nr", "company", "location", "contact_type", "name", "role",
        "email", "email_status", "sprache", "ansatz", "icp_prio",
        "mouaz_kontakt", "betreff", "anhang", "teaser_pfad",
        "versand_status", "template_ref", "alt_email", "central_email",
        "notizen", "notizen_extra", "email_body",
    ]

    wb = openpyxl.load_workbook(XLSX_PATH)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    wrap = Alignment(wrap_text=True, vertical="top")
    for r_idx, r in enumerate(rows_out, 2):
        for c_idx, h in enumerate(headers, 1):
            val = r.get(h, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap

    widths = {
        "company": 36, "name": 22, "role": 32, "email": 34,
        "betreff": 44, "email_body": 72, "mouaz_kontakt": 32,
        "notizen": 28, "teaser_pfad": 42,
    }
    for c_idx, h in enumerate(headers, 1):
        letter = get_column_letter(c_idx)
        ws.column_dimensions[letter].width = widths.get(h, 14)

    ws.freeze_panes = "A2"
    out_path = XLSX_PATH
    try:
        wb.save(out_path)
    except PermissionError:
        out_path = BASE / "260609_CRM_Funnel_Interpack_List_Email.xlsx"
        wb.save(out_path)
        print(
            f"HINWEIS: Original gesperrt (Excel offen?). "
            f"Gespeichert als: {out_path}"
        )
    print(f"Written {len(rows_out)} rows to sheet '{SHEET_NAME}' in {out_path}")


if __name__ == "__main__":
    main()
