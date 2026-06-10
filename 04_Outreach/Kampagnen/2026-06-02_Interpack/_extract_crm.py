"""Extract and filter CRM Funnel Interpack list."""
import json
import re
from openpyxl import load_workbook
from pathlib import Path

XLSX = Path(
    r"C:\Users\olafp\Desktop\Arbeitsordner\260516_LFL_Marketing_Teaser"
    r"\04_Outreach\Kampagnen\2026-06-02_Interpack\260605_CRM_Funnel_Interpack_List.xlsx"
)
OUT_DIR = XLSX.parent


def parse_headcount(val) -> tuple[int | None, str]:
    if val is None:
        return None, ""
    raw = str(val).strip()
    # take first number found; handle >3.000, ~1.373, 6,900
    cleaned = raw.replace(".", "").replace(",", "")
    nums = re.findall(r"\d+", cleaned)
    if not nums:
        return None, raw
    # for ranges like 1373;4300 take first
    n = int(nums[0])
    return n, raw


def is_prio_a(remarks: str) -> bool:
    if not remarks:
        return False
    r = remarks.lower()
    # "Outreach-Prio lt. InMail-Liste: A" or "md-Prio A" or "Prio A"
    if re.search(r"(?:inmail-liste|outreach-prio|md-prio|prio)\s*[^a-z]*\ba\b", r):
        return True
    if re.search(r":\s*a\.", r):  # ": A. Große"
        return True
    return False


def row_to_dict(header, row):
    d = {}
    for i, h in enumerate(header):
        if h:
            key = str(h).split("\n")[0].strip()
            d[key] = row[i] if i < len(row) else None
    return d


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Tabelle2"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = rows[0]
    all_rows = []
    filtered = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        d = row_to_dict(header, row)
        hc, hc_raw = parse_headcount(d.get("Headcount (Est.)"))
        d["_headcount_parsed"] = hc
        d["_headcount_raw"] = hc_raw
        d["_prio_a"] = is_prio_a(str(d.get("Remarks") or ""))
        all_rows.append(d)
        if d["_prio_a"] and hc is not None and hc > 200:
            filtered.append(d)

    (OUT_DIR / "_crm_all.json").write_text(
        json.dumps(all_rows, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    (OUT_DIR / "_filtered_prio_a_gt200.json").write_text(
        json.dumps(filtered, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(f"Total: {len(all_rows)}, Filtered: {len(filtered)}")
    for f in filtered:
        print(f"  - {f['Company Name (Full legal company name)']} | MA={f['_headcount_parsed']}")


if __name__ == "__main__":
    main()
