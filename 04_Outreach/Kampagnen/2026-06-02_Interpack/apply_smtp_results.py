#!/usr/bin/env python3
"""Apply SMTP results: update emails, smtp_check, email_routing_info@."""
import importlib.util
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

BASE = Path(__file__).resolve().parent
XLSX = BASE / "260605_CRM_Funnel_Interpack_List.xlsx"
SHEET = "Email_Outreach_Interpack_Warm"
RESULTS = BASE / "_smtp_results.json"

spec = importlib.util.spec_from_file_location("build", BASE / "_build_email_outreach_sheet.py")
BUILD = importlib.util.module_from_spec(spec)
spec.loader.exec_module(BUILD)


def language_is_de(lang) -> bool:
    if not lang:
        return True
    return str(lang).strip().lower().startswith("deu")


def classify_fail(detail: str) -> str:
    d = detail.lower()
    if "user unknown" in d or "does not exist" in d or "5.1.1" in d:
        return "ungültig"
    if "access denied" in d or "5.4.1" in d:
        return "SMTP blockiert (Adresse evtl. OK)"
    if "spamhaus" in d or "blacklisted" in d or "5.7.1" in d:
        return "SMTP nicht prüfbar (IP blockiert)"
    return "fail"


def routing_text(row, meta, de, central):
    name = row["name"]
    role = row["role"]
    body = row["email_body"] or ""
    short = meta.get("short", row["company"])

    if de:
        header = (
            f"An: {central}\n"
            f"Betreff: Bitte weiterleiten an {name.split()[-1]} — Nach interpack ({short})\n\n"
            f"Guten Tag,\n\n"
            f"könnten Sie diese Nachricht bitte an {name} ({role}) weiterleiten?\n\n"
            f"---\n\n"
        )
        footer = (
            f"\n\n---\n\nVielen Dank für die Weiterleitung.\n\n"
            f"Beste Grüße\nOlaf Pick\nLoopForgeLab GbR\nolaf@loopforgelab.com"
        )
    else:
        header = (
            f"To: {central}\n"
            f"Subject: Please forward to {name.split()[-1]} — follow-up after interpack ({short})\n\n"
            f"Dear Sir or Madam,\n\n"
            f"could you please forward this message to {name} ({role})?\n\n"
            f"---\n\n"
        )
        footer = (
            f"\n\n---\n\nThank you for forwarding.\n\n"
            f"Best regards\nOlaf Pick\nLoopForgeLab GbR\nolaf@loopforgelab.com"
        )

    core = body
    for cut in (
        "Dear Sir or Madam", "Guten Tag,", "could you please forward",
        "könnten Sie diese Nachricht", "Vielen Dank", "Thank you for forwarding",
        "Beste Grüße", "Best regards",
    ):
        if cut in core:
            parts = core.split(cut, 1)
            if cut in ("Dear Sir or Madam", "Guten Tag,"):
                core = parts[-1].strip()
            else:
                core = parts[0].strip()
    core = re.sub(r"^---+\\s*", "", core).strip()
    return f"{header}{core}{footer}"


# Manual overrides from SMTP test
DIRECT_OK = {
    "Fredrik Sandmark": "fredrik.sandmark@beumer.com",
    "Christopher Kirsch": "christopher.kirsch@beumergroup.com",
}


def main():
    data = json.loads(RESULTS.read_text(encoding="utf-8"))
    by_row = {d["row"]: d for d in data}

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]

    def col(name):
        return headers.index(name)

    lang_i = next(i for i, h in enumerate(headers) if h and "language" in str(h).lower())
    smtp_i = headers.index("smtp_check") if "smtp_check" in headers else None
    route_i = headers.index("email_routing_info@") if "email_routing_info@" in headers else None

    if smtp_i is None:
        headers.append("smtp_check")
        ws.cell(1, len(headers), "smtp_check")
        smtp_i = len(headers) - 1
    if route_i is None:
        headers.append("email_routing_info@")
        ws.cell(1, len(headers), "email_routing_info@")
        route_i = len(headers) - 1

    wrap = Alignment(wrap_text=True, vertical="top")
    c = {k: col(k) for k in [
        "nr", "name", "company", "role", "contact_type", "email", "email_status",
        "central_email", "email_body",
    ]}

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, c["name"] + 1).value
        if name in (None, "k.A."):
            continue

        row_data = by_row.get(r)
        if not row_data:
            continue

        nr = int(ws.cell(r, c["nr"] + 1).value)
        meta = BUILD.COMPANY_META.get(nr, {"short": ws.cell(r, c["company"] + 1).value})
        de = language_is_de(ws.cell(r, lang_i + 1).value)
        central = ws.cell(r, c["central_email"] + 1).value or "info@company.com"
        row = {k: ws.cell(r, c[k] + 1).value for k in c}

        tested = {t["email"]: t for t in row_data["tested"]}
        primary = row_data["primary"]

        # Beumer: verified direct
        if name in DIRECT_OK:
            ok_email = DIRECT_OK[name]
            ws.cell(r, c["email"] + 1, ok_email)
            ws.cell(r, smtp_i + 1, f"smtp OK — {ok_email}")
            ws.cell(r, c["email_status"] + 1, "smtp verifiziert")
            ws.cell(r, route_i + 1, "")
            # Restore direct mail body (not info@ routing wrapper)
            ansatz = ws.cell(r, headers.index("ansatz") + 1).value
            fake = {"name": name, "contact_type": row["contact_type"]}
            body = BUILD.body_a(fake, meta, de) if ansatz == "A" else BUILD.body_b(fake, meta, de)
            ws.cell(r, c["email_body"] + 1, body).alignment = wrap
            betreff = BUILD.betreff(meta["short"], row["contact_type"], de, ansatz)
            ws.cell(r, headers.index("betreff") + 1, betreff)
            ws.cell(r, headers.index("versand_status") + 1, "vorbereitet")
            continue

        prim = tested.get(primary, {})
        working_direct = [
            e for e, t in tested.items()
            if e != primary and t["status"] == "ok"
            and not e.startswith("press@")
            and not e.startswith("presse@")
            and not e.startswith("pr@")
            and "agentur" not in e
            and e.split("@")[0] not in ("info", "sales", "personal", "tablet", "karriere")
        ]

        if prim.get("status") == "ok":
            ws.cell(r, smtp_i + 1, f"smtp OK — {primary}")
            ws.cell(r, route_i + 1, "")
            continue

        # Primary failed
        kind = classify_fail(prim.get("detail", ""))
        failed_alts = [e for e, t in tested.items() if t["status"] == "fail" and e != primary]

        if kind == "SMTP nicht prüfbar (IP blockiert)":
            smtp_label = f"smtp UNPRÜFBAR — {primary} (IP blockiert, manuell testen)"
            needs_routing = True
        elif kind == "SMTP blockiert (Adresse evtl. OK)":
            smtp_label = f"smtp UNKLAR — {primary} (externe Probe blockiert)"
            needs_routing = True
        else:
            smtp_label = f"smtp FAIL — {primary} ({kind})"
            needs_routing = True

        if failed_alts:
            smtp_label += f" | fail auch: {', '.join(failed_alts[:3])}"

        ws.cell(r, smtp_i + 1, smtp_label)
        old_st = ws.cell(r, c["email_status"] + 1).value or ""
        if "smtp" not in str(old_st).lower():
            ws.cell(r, c["email_status"] + 1, f"{old_st} | {smtp_label}".strip(" |"))

        if needs_routing:
            rt = routing_text(row, meta, de, central)
            ws.cell(r, route_i + 1, rt).alignment = wrap
            ws.cell(r, headers.index("versand_status") + 1, "routing vorbereitet")

    wb.save(XLSX)
    print("Applied SMTP results to", XLSX)


if __name__ == "__main__":
    main()
