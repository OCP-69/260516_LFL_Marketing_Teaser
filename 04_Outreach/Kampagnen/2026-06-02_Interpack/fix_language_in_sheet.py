#!/usr/bin/env python3
"""Sync email_body/betreff to Language column in CRM outreach sheet."""
import importlib.util
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

BASE = Path(__file__).resolve().parent
XLSX = BASE / "260605_CRM_Funnel_Interpack_List.xlsx"
SHEET = "Email_Outreach_Interpack_Warm"

# Load build module
spec = importlib.util.spec_from_file_location(
    "build", BASE / "_build_email_outreach_sheet.py"
)
build = importlib.util.module_from_spec(spec)
spec.loader.exec_module(build)

DE_MARKERS = (
    "Guten Tag", "Mein Kollege", "Mouaz war", "Beste Grüße",
    "Pilot Q2/2026 · kein ERP", "Nach interpack —", "interpack-Follow-up",
    "Limitiert ", "Gehören teure", "Ist \"Kosten",
)
EN_MARKERS = ("Dear ", "Best regards", "My colleague", "Mouaz spoke", "After interpack")


def is_german_text(text: str) -> bool:
    if not text or text.startswith("—"):
        return False
    return any(m in text for m in DE_MARKERS)


def is_english_text(text: str) -> bool:
    if not text or text.startswith("—"):
        return False
    return any(m in text for m in EN_MARKERS)


def language_is_de(lang) -> bool:
    if not lang:
        return True
    s = str(lang).strip().lower()
    return s.startswith("deu") or s == "de"


def find_col(headers, *candidates):
    for i, h in enumerate(headers):
        if not h:
            continue
        hs = str(h).lower()
        for c in candidates:
            if c.lower() in hs:
                return i
    return None


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]

    col = {
        "nr": headers.index("nr"),
        "name": headers.index("name"),
        "contact_type": headers.index("contact_type"),
        "language": find_col(headers, "language"),
        "sprache": headers.index("sprache"),
        "ansatz": headers.index("ansatz"),
        "betreff": headers.index("betreff"),
        "mouaz": headers.index("mouaz_kontakt"),
        "body": headers.index("email_body"),
        "template": headers.index("template_ref"),
    }

    wrap = Alignment(wrap_text=True, vertical="top")
    fixed = []
    ok = 0

    for r in range(2, ws.max_row + 1):
        nr_val = ws.cell(r, col["nr"] + 1).value
        if nr_val is None:
            continue
        nr = int(nr_val)
        lang_cell = ws.cell(r, col["language"] + 1).value
        ansatz = ws.cell(r, col["ansatz"] + 1).value
        body = ws.cell(r, col["body"] + 1).value or ""

        if ansatz == "AUSSETZEN" or not lang_cell:
            continue

        de = language_is_de(lang_cell)
        row = {
            "name": ws.cell(r, col["name"] + 1).value,
            "contact_type": ws.cell(r, col["contact_type"] + 1).value,
        }
        meta = build.COMPANY_META.get(nr, {})
        short = meta.get("short", "")

        needs_fix = (de and is_english_text(body)) or (not de and is_german_text(body))
        if not needs_fix:
            ok += 1
            continue

        if ansatz == "A":
            new_body = build.body_a(row, meta, de)
            new_subj = build.betreff(short, row["contact_type"], de, "A")
        else:
            new_body = build.body_b(row, meta, de)
            new_subj = build.betreff(short, row["contact_type"], de, "B")

        new_mouaz = meta.get("mouaz_de" if de else "mouaz_en", "")
        new_sprache = "DE" if de else "EN"
        new_tpl = f"Ansatz_{ansatz}_kurz_{new_sprache}"

        ws.cell(r, col["betreff"] + 1, value=new_subj).alignment = wrap
        ws.cell(r, col["body"] + 1, value=new_body).alignment = wrap
        ws.cell(r, col["mouaz"] + 1, value=new_mouaz).alignment = wrap
        ws.cell(r, col["sprache"] + 1, value=new_sprache)
        ws.cell(r, col["template"] + 1, value=new_tpl)

        fixed.append({
            "row": r,
            "nr": nr,
            "name": row["name"],
            "language": lang_cell,
            "betreff": new_subj,
        })

    try:
        wb.save(XLSX)
    except PermissionError:
        alt = BASE / "260609_CRM_Funnel_Interpack_List_Email.xlsx"
        wb.save(alt)
        print(f"Original gesperrt — gespeichert als: {alt}")
        sys.exit(0)

    print(f"OK (bereits passend): {ok}")
    print(f"Aktualisiert: {len(fixed)}")
    for f in fixed:
        print(f"  Zeile {f['row']}: nr={f['nr']} {f['name']} ({f['language']})")
        print(f"    Betreff: {f['betreff']}")


if __name__ == "__main__":
    main()
