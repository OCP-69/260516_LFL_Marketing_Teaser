#!/usr/bin/env python3
"""Rewrite email_body + betreff with clearer narrative structure."""
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

BASE = Path(__file__).resolve().parent
XLSX = BASE / "260605_CRM_Funnel_Interpack_List.xlsx"
SHEET = "Email_Outreach_Interpack_Warm"
CALENDLY = "https://calendly.com/olaf-pick/olaf-15min"

# assumption + topics per company (CEO vs Engineering)
META = {
    1: {
        "short": "Beumer Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "Variantenkalkulation bei Intralogistik- und Sonderprojekten",
        "messe_en": "variant costing on intralogistics and custom projects",
        "assume_ceo_de": "dass bei großen Intralogistik-Projekten die Engineering-Kapazität für Angebotsvarianten knapper ist als die Nachfrage nach schnellen RFQs",
        "assume_eng_de": "dass teure Änderungen nach Freigabe in Ihrem Innovationsumfeld eher die Ausnahme sind — oder doch zum Alltag gehören",
        "assume_ceo_en": "that engineering capacity for quote variants is tighter than customer demand for fast RFQs on large intralogistics projects",
        "assume_eng_en": "that costly post-release changes in your innovation environment are the exception — or part of everyday work",
        "topics_ceo_de": "mehr Angebotsvarianten bei gleicher Engineering-Kapazität und belastbarere Kosten aus BOM/CAD bereits im Design",
        "topics_eng_de": "Varianten aus BOM, CAD und Einkauf vor Freigabe in Euro — ohne neues ERP",
        "topics_ceo_en": "more quote variants with the same engineering capacity and reliable BOM/CAD costs at design stage",
        "topics_eng_en": "BOM, CAD and purchasing variants in euros before release — without a new ERP",
    },
    2: {
        "short": "Handtmann Processing",
        "mouaz_de": "Kollegen von Ihrem Messestand",
        "mouaz_en": "colleagues at your stand",
        "messe_de": "Varianten und Kosten vor der Freigabe bei Processing-Lösungen",
        "messe_en": "variants and costs before release on processing solutions",
        "assume_ceo_de": "dass bei kundenspezifischen Varianten die Marge erst dann belastbar wird, wenn Konstruktion, Einkauf und Kalkulation mehrere Runden drehen",
        "assume_eng_de": "dass Material- und Prozess-What-ifs in der Konstruktion oft erst spät mit belastbaren Euro-Zahlen unterlegt werden",
        "assume_ceo_en": "that margin on custom variants only becomes reliable after several rounds between design, purchasing and costing",
        "assume_eng_en": "that material and process what-ifs in design are often backed by reliable euro figures only late in the process",
        "topics_ceo_de": "Kostensicherheit vor Freigabe, weniger teure Nacharbeit und klarere Margen pro Auftrag",
        "topics_eng_de": "BOM-/CAD-Varianten im Design durchrechnen und Freigaben mit belastbaren Zahlen untermauern",
        "topics_ceo_en": "cost certainty before release, less costly rework and clearer margins per order",
        "topics_eng_en": "costing BOM/CAD variants at design stage and backing release decisions with reliable figures",
    },
    3: {
        "short": "Winkler+Dünnebier",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "RFQ-Tempo und Variantenkalkulation bei Süßwaren-Sondermaschinen",
        "messe_en": "RFQ speed and variant costing on confectionery custom machinery",
        "assume_ceo_de": "dass Sonderanfragen nach der Messe in klassischen Excel-Schleifen landen, obwohl BOM- und CAD-Daten schon vorliegen",
        "assume_eng_de": "dass Stücklisten- und CAD-Varianten vor Freigabe noch zu oft ohne belastbare Kostenentscheidung bleiben",
        "assume_ceo_en": "that custom inquiries after the show still end up in spreadsheet loops although BOM and CAD data already exist",
        "assume_eng_en": "that BOM and CAD variants too often remain without a reliable cost decision before release",
        "topics_ceo_de": "schnellere RFQs und Varianten im Design statt Wochen in manuellen Kalkulationsschleifen",
        "topics_eng_de": "Varianten vor Freigabe in Euro und weniger Überraschungen in der Fertigung",
        "topics_ceo_en": "faster RFQs and variants at design instead of weeks in manual costing loops",
        "topics_eng_en": "variants in euros before release and fewer surprises in production",
    },
    4: {
        "short": "KÖRA-PACKMAT",
        "mouaz_de": "Herrn Faller (Einkauf & Materialwirtschaft)",
        "mouaz_en": "Mr Faller (purchasing and materials)",
        "messe_de": "Lieferanten- und Materialszenarien im Maschinenbau",
        "messe_en": "supplier and material scenarios in machinery engineering",
        "assume_ceo_de": "dass bei kundenspezifischen Maschinen Lieferanten- und Materialentscheidungen im Design noch zu wenig in die Angebotsbasis einfließen",
        "assume_eng_de": "dass Einkaufspreise und BOM-Alternativen in der Konstruktion oft erst nach Rückfragen an Einkauf belastbar werden",
        "assume_ceo_en": "that supplier and material decisions at design stage still feed too little into the quote basis on custom machinery",
        "assume_eng_en": "that purchase prices and BOM alternatives in design often become reliable only after back-and-forths with purchasing",
        "topics_ceo_de": "Materialszenarien und Lieferantenrisiken schon im Design sichtbar machen",
        "topics_eng_de": "Einkaufspreise und BOM-Alternativen im Konstruktionsmoment vergleichen — ohne ERP-Projekt",
        "topics_ceo_en": "making material scenarios and supplier risks visible at design stage",
        "topics_eng_en": "comparing purchase prices and BOM alternatives during design — without an ERP project",
    },
    5: {
        "short": "Syntegon",
        "mouaz_de": "Herrn Ruener von Ihrem Team",
        "mouaz_en": "Mr Ruener from your team",
        "messe_de": "Variantenflut nach der Messe und Engineering-Kapazität",
        "messe_en": "post-show variant volume and engineering capacity",
        "assume_ceo_de": "dass nach interpack die Anzahl der zu bewertenden Varianten steigt, während die Engineering-Kapazität für Angebote begrenzt bleibt",
        "assume_eng_de": "dass Sekundärverpackungs-Varianten im Design noch zu viele Freigabe-Runden mit Vertrieb auslösen",
        "assume_ceo_en": "that after interpack the number of variants to evaluate rises while engineering capacity for quotes stays limited",
        "assume_eng_en": "that secondary packaging variants at design still trigger too many release rounds with sales",
        "topics_ceo_de": "Variantenflut schneller bewerten und Engineering-Kapazität für Angebote statt Nachkalkulation nutzen",
        "topics_eng_de": "Varianten im Design in Kosten übersetzen und Freigaben entlasten",
        "topics_ceo_en": "evaluating variant floods faster and using engineering for quotes, not rework costing",
        "topics_eng_en": "turning variants into costs at design and easing release workflows",
    },
    6: {
        "short": "Schubert Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "modulare Verpackungslinien und Variantenkalkulation",
        "messe_en": "modular packaging lines and variant costing",
        "assume_ceo_de": "dass modulare Linien zwar schnell konfiguriert, aber in Angeboten noch zu langsam in belastbare Kosten überführt werden",
        "assume_eng_de": "dass Pick-and-place- und Modul-Varianten pro RFQ-Schleife noch zu viel manuelle Kalkulationsarbeit binden",
        "assume_ceo_en": "that modular lines are configured quickly but still translated too slowly into reliable costs in quotes",
        "assume_eng_en": "that pick-and-place and module variants still tie up too much manual costing work per RFQ loop",
        "topics_ceo_de": "modulare Linien schneller kalkulieren aus bestehenden BOM- und CAD-Daten",
        "topics_eng_de": "RFQ-Varianten in Tagen statt Wochen — ohne ERP-Ersatz",
        "topics_ceo_en": "costing modular lines faster from existing BOM and CAD data",
        "topics_eng_en": "RFQ variants in days not weeks — without replacing your ERP",
    },
    7: {
        "short": "SMI Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "RFQ speed on PET and packaging lines",
        "messe_en": "RFQ speed on PET and packaging lines",
        "assume_ceo_de": "that quote preparation on PET/packaging lines still takes weeks while customers expect faster decisions",
        "assume_eng_de": "that RFQ variants for packaging lines still require too many manual iterations between BOM, CAD and costing",
        "assume_ceo_en": "that quote preparation on PET/packaging lines still takes weeks while customers expect faster decisions",
        "assume_eng_en": "that RFQ variants for packaging lines still require too many manual iterations between BOM, CAD and costing",
        "topics_ceo_de": "faster quotes and costs from BOM and CAD already in engineering",
        "topics_eng_de": "RFQ variants straight from BOM and CAD — pilot without native CAD integration",
        "topics_ceo_en": "faster quotes and costs from BOM and CAD already in engineering",
        "topics_eng_en": "RFQ variants straight from BOM and CAD — pilot without native CAD integration",
    },
    8: {
        "short": "Romaco Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "Pharma-Verpackungsprojekte und Varianten",
        "messe_en": "pharma packaging projects and variants",
        "assume_ceo_de": "dass bei Pharma-Projektvarianten Angebot und Marge erst spät belastbar werden, wenn mehrere Abteilungen eingebunden sind",
        "assume_eng_de": "dass compliance-relevante Änderungen im Design oft ohne sofortige Kostensicht bewertet werden",
        "assume_ceo_en": "that on pharma project variants quote and margin become reliable only late, once several departments are involved",
        "assume_eng_en": "that compliance-relevant design changes are often assessed without immediate cost visibility",
        "topics_ceo_de": "Varianten früher in belastbare Kosten und weniger Risiko in Angebot und Marge",
        "topics_eng_de": "Varianten im Design durchrechnen — Pilot startet mit wenigen Stücklisten",
        "topics_ceo_en": "reliable costs on variants earlier and less risk in quote and margin",
        "topics_eng_en": "costing variants at design — pilot starts with a few BOMs",
    },
    9: {
        "short": "MULTIVAC",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "integrierte Verpackungslösungen und RFQ-Last",
        "messe_en": "integrated packaging solutions and RFQ workload",
        "assume_ceo_de": "dass integrierte Lösungen zwar technisch überzeugen, in der Angebotsphase aber noch zu viel Engineering für Variantenkalkulation binden",
        "assume_eng_de": "dass Linien- und Modulvarianten aus BOM, CAD und Einkauf noch nicht in einer gemeinsamen Kostensicht zusammenlaufen",
        "assume_ceo_en": "that integrated solutions convince technically but still tie up too much engineering for variant costing in the quote phase",
        "assume_eng_en": "that line and module variants from BOM, CAD and purchasing do not yet converge in one cost view",
        "topics_ceo_de": "RFQ-Tempo erhöhen und Engineering für Varianten entlasten",
        "topics_eng_de": "eine Kostensicht über BOM, CAD und Einkauf — ohne ERP-Rollout",
        "topics_ceo_en": "increase RFQ speed and free up engineering for variants",
        "topics_eng_en": "one cost view across BOM, CAD and purchasing — without an ERP rollout",
    },
    10: {
        "short": "Sidel",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "equipment portfolio and customer specifications",
        "messe_en": "equipment portfolio and customer specifications",
        "assume_ceo_de": "that portfolio variants are evaluated faster technically than economically at design stage",
        "assume_eng_de": "that customer specs and portfolio variants in engineering still produce costly rework after release",
        "assume_ceo_en": "that portfolio variants are evaluated faster technically than economically at design stage",
        "assume_eng_en": "that customer specs and portfolio variants in engineering still produce costly rework after release",
        "topics_ceo_de": "faster economic evaluation of portfolio variants at design stage",
        "topics_eng_de": "customer specs in euros during engineering — less rework after release",
        "topics_ceo_en": "faster economic evaluation of portfolio variants at design stage",
        "topics_eng_en": "customer specs in euros during engineering — less rework after release",
    },
    11: {
        "short": "Krones",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "Abfüll- und Verpackungstechnik nach der Messe",
        "messe_en": "filling and packaging technology after the show",
        "assume_ceo_de": "dass bei Linienprojekten die Angebotskapazität eher durch Variantenkalkulation limitiert wird als durch den Vertrieb",
        "assume_eng_de": "dass Linienvarianten vor Freigabe noch zu viele Iterationen mit Vertrieb und Einkauf auslösen",
        "assume_ceo_en": "that on line projects quote capacity is limited more by variant costing than by sales",
        "assume_eng_en": "that line variants before release still trigger too many iterations with sales and purchasing",
        "topics_ceo_de": "Varianten früher in belastbare Kosten und mehr Angebotskapazität",
        "topics_eng_de": "BOM-/CAD-Varianten vor Freigabe — weniger manuelle Schleifen",
        "topics_ceo_en": "reliable costs on variants earlier and more quote capacity",
        "topics_eng_en": "BOM/CAD variants before release — fewer manual loops",
    },
    12: {
        "short": "KHS",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "Linienengineering und Angebotsgeschwindigkeit",
        "messe_en": "line engineering and quote speed",
        "assume_ceo_de": "dass Linienengineering in Angebote übersetzt wird, bevor Variantenkosten im Design wirklich belastbar sind",
        "assume_eng_de": "dass Cross-Function-Varianten aus Stückliste und CAD noch zu oft manuell nachkalkuliert werden",
        "assume_ceo_en": "that line engineering is turned into quotes before variant costs are truly reliable at design stage",
        "assume_eng_en": "that cross-function variants from BOM and CAD are still manually re-costed too often",
        "topics_ceo_de": "Variantenkosten im Design statt erst im Nachgang — schnellere Angebote",
        "topics_eng_de": "Varianten sofort kalkulieren — Pilot mit wenigen BOMs",
        "topics_ceo_en": "variant costs at design, not after the fact — faster quotes",
        "topics_eng_en": "cost variants instantly — pilot with a few BOMs",
    },
    13: {
        "short": "IMA Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "packaging machinery and digitalisation",
        "messe_en": "packaging machinery and digitalisation",
        "assume_ceo_de": "that packaging machine variants still take too long to turn into quotes without a major IT project",
        "assume_eng_de": "that Industry 4.0 data (BOM, CAD, purchasing) is not yet used systematically for cost scenarios at design",
        "assume_ceo_en": "that packaging machine variants still take too long to turn into quotes without a major IT project",
        "assume_eng_en": "that Industry 4.0 data (BOM, CAD, purchasing) is not yet used systematically for cost scenarios at design",
        "topics_ceo_de": "faster quotes and digitalisation without ERP rollout",
        "topics_eng_de": "cost scenarios at design from existing data — GDPR, EU hosting",
        "topics_ceo_en": "faster quotes and digitalisation without ERP rollout",
        "topics_eng_en": "cost scenarios at design from existing data — GDPR, EU hosting",
    },
    14: {
        "short": "Harro Höfliger",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "Sonderanlagen und Freigabe-Prozesse",
        "messe_en": "special-purpose machines and release processes",
        "assume_ceo_de": "dass bei Sonderanlagen die Marge erst dann klar wird, wenn Varianten und Kosten nach der Freigabe nachjustiert werden",
        "assume_eng_de": "dass Freigabe-Entscheidungen noch zu oft ohne belastbare Euro-Zahlen aus BOM und CAD getroffen werden",
        "assume_ceo_en": "that on special-purpose machines margin only becomes clear when variants and costs are adjusted after release",
        "assume_eng_en": "that release decisions are still too often taken without reliable euro figures from BOM and CAD",
        "topics_ceo_de": "Kosten vor Freigabe belastbar machen und Marge-Risiko bei Sonderprojekten senken",
        "topics_eng_de": "Varianten im Design mit belastbaren Zahlen untermauern",
        "topics_ceo_en": "reliable costs before release and lower margin risk on custom projects",
        "topics_eng_en": "back variants at design with reliable figures",
    },
    15: {
        "short": "GEA Group",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "Prozesstechnik und Angebotsvarianten",
        "messe_en": "process technology and quote variants",
        "assume_ceo_de": "dass Engineering-Kapazität bei Prozesstechnik-Varianten noch zu stark in Nachkalkulation statt in neue Angebote fließt",
        "assume_eng_de": "dass Angebotsvarianten aus Stücklisten und Einkaufsdaten noch nicht in einer durchgängigen Kostensicht zusammenlaufen",
        "assume_ceo_en": "that engineering capacity on process technology variants still flows too much into rework costing instead of new quotes",
        "assume_eng_en": "that quote variants from BOMs and purchasing data do not yet converge in one continuous cost view",
        "topics_ceo_de": "Varianten schneller in Angebote und Kapazität für Innovation freihalten",
        "topics_eng_de": "Intelligenzschicht auf bestehende Daten — kein ERP-Ersatz",
        "topics_ceo_en": "variants into quotes faster and capacity for innovation",
        "topics_eng_en": "intelligence layer on existing data — not an ERP replacement",
    },
    16: {
        "short": "Fette Compacting",
        "mouaz_de": "Kollegen von Ihrem Messestand",
        "mouaz_en": "colleagues at your stand",
        "messe_de": "Tablettierung und Variantenkosten",
        "messe_en": "tableting and variant costs",
        "assume_ceo_de": "dass Tablettier-Varianten wirtschaftlich oft erst nach der Freigabe vollständig durchsichtig werden",
        "assume_eng_de": "dass Maschinen- und Werkzeugvarianten im Design noch zu wenig mit belastbaren Kosten verknüpft sind",
        "assume_ceo_en": "that tableting variants economically often become fully transparent only after release",
        "assume_eng_en": "that machine and tooling variants at design are still too weakly linked to reliable costs",
        "topics_ceo_de": "Varianten früher in Kosten und weniger Überraschung nach Freigabe",
        "topics_eng_de": "Varianten im Design durchrechnen — Pilot Q2/2026",
        "topics_ceo_en": "variants in costs earlier and fewer surprises after release",
        "topics_eng_en": "cost variants at design — Q2/2026 pilot",
    },
    18: {
        "short": "TOMRA",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "food sorting equipment and project variants",
        "messe_en": "food sorting equipment and project variants",
        "assume_ceo_de": "that variant costs on food sorting projects become visible too late for fast quote decisions",
        "assume_eng_de": "that equipment variants from BOM and engineering data still require manual costing loops",
        "assume_ceo_en": "that variant costs on food sorting projects become visible too late for fast quote decisions",
        "assume_eng_en": "that equipment variants from BOM and engineering data still require manual costing loops",
        "topics_ceo_de": "variant costs earlier and faster quote decisions",
        "topics_eng_de": "cost equipment variants from existing data — without ERP rollout",
        "topics_ceo_en": "variant costs earlier and faster quote decisions",
        "topics_eng_en": "cost equipment variants from existing data — without ERP rollout",
    },
    19: {
        "short": "Uhlmann Pac-Systeme",
        "mouaz_de": "Kollegen von Ihrem interpack-Stand",
        "mouaz_en": "colleagues at your interpack stand",
        "messe_de": "Pharma-Verpackungslinien und Varianten",
        "messe_en": "pharma packaging lines and variants",
        "assume_ceo_de": "dass bei Pharma-Linien Varianten und Margen vor Freigabe noch zu unscharf bleiben",
        "assume_eng_de": "dass Linien- und Modulvarianten im Solution Engineering noch zu viele manuelle Kalkulationsschritte brauchen",
        "assume_ceo_en": "that on pharma lines variants and margins remain too unclear before release",
        "assume_eng_en": "that line and module variants in solution engineering still need too many manual costing steps",
        "topics_ceo_de": "Varianten und Margen vor Freigabe klarer — weniger teure Projektänderungen",
        "topics_eng_de": "Varianten direkt in Euro im Engineering — DSGVO, Daten beim Kunden",
        "topics_ceo_en": "clearer variants and margins before release — fewer costly project changes",
        "topics_eng_en": "variants in euros during engineering — GDPR, data stays on your side",
    },
}


def salutation_de(name: str) -> str:
    if not name or name == "k.A.":
        return "Guten Tag,"
    if "Jessica" in name:
        return f"Guten Tag Frau {name.split()[-1]},"
    if name.startswith("Dr."):
        return f"Guten Tag Herr Dr. {name.split()[-1]},"
    return f"Guten Tag Herr {name.split()[-1]},"


def salutation_en(name: str) -> str:
    if not name or name == "k.A.":
        return "Dear Sir or Madam,"
    last = name.split()[-1]
    if "Jessica" in name:
        return f"Dear Ms {last},"
    if name.startswith("Dr."):
        return f"Dear Dr {last},"
    if "Andersen" in name:
        return f"Dear Ms {last},"
    return f"Dear Mr {last},"


def language_is_de(lang) -> bool:
    if not lang:
        return True
    return str(lang).strip().lower().startswith("deu")


def build_email(row: dict, meta: dict, de: bool, ansatz: str) -> tuple[str, str]:
    short = meta["short"]
    is_ceo = row["contact_type"] == "CEO/GF"
    sal = salutation_de(row["name"]) if de else salutation_en(row["name"])
    mouaz = meta["mouaz_de"] if de else meta["mouaz_en"]
    messe = meta["messe_de"] if de else meta["messe_en"]
    assume = meta["assume_ceo_de" if is_ceo else "assume_eng_de"] if de else meta[
        "assume_ceo_en" if is_ceo else "assume_eng_en"
    ]
    topics = meta["topics_ceo_de" if is_ceo else "topics_eng_de"] if de else meta[
        "topics_ceo_en" if is_ceo else "topics_eng_en"
    ]
    sig_de = "Beste Grüße\nOlaf Pick\nLoopForgeLab GbR\nolaf@loopforgelab.com"
    sig_en = "Best regards\nOlaf Pick\nLoopForgeLab GbR\nolaf@loopforgelab.com"

    if de:
        betreff = (
            f"Nach interpack — passt das Thema zu {short}?"
            if ansatz == "A"
            else f"interpack-Follow-up — Variantenkosten bei {short}"
        )
        p1 = (
            f"im Gespräch mit Maschinenbauern im Verpackungs- und Sondermaschinenbau hören wir häufig, "
            f"{assume}. Ist diese Annahme für {short} zutreffend — oder sehen Sie das deutlich anders?"
        )
        p2 = (
            f"Auf der interpack war mein Kollege Mouaz Al-Qudsi (CTO, LoopForgeLab) mit {mouaz} "
            f"im Austausch. Im Gespräch ging es unter anderem um {messe}."
        )
        p3 = (
            f"Wir entwickeln mit Forge Engine eine Kostensicht aus BOM, CAD und Einkauf — "
            f"ohne ERP-Ersatz, als Pilot mit wenigen Stücklisten. "
            f"Wären Themen wie {topics} für {short} 2026 relevant — oder eher nicht?"
        )
        p4 = (
            f"Falls Sie das grob einordnen möchten, können wir Details gern in einem kurzen "
            f"Gespräch (15 Min.) klären — {CALENDLY} — oder Sie antworten einfach auf diese Mail."
        )
        body = f"{sal}\n\n{p1}\n\n{p2}\n\n{p3}\n\n{p4}\n\n{sig_de}"
    else:
        betreff = (
            f"After interpack — does this fit {short}?"
            if ansatz == "A"
            else f"interpack follow-up — variant costing at {short}"
        )
        p1 = (
            f"In conversations with machinery OEMs in packaging and custom engineering we often hear "
            f"{assume}. Does this assumption fit {short} — or do you see it very differently?"
        )
        p2 = (
            f"At interpack my colleague Mouaz Al-Qudsi (CTO, LoopForgeLab) spoke with {mouaz}. "
            f"The conversation touched on {messe}."
        )
        p3 = (
            f"We are building Forge Engine as a cost view from BOM, CAD and purchasing — "
            f"not an ERP replacement, starting as a pilot with a few BOMs. "
            f"Would topics such as {topics} be relevant for {short} in 2026 — or rather not?"
        )
        p4 = (
            f"If you would like to sense-check this, we can clarify details in a "
            f"15-minute call — {CALENDLY} — or simply reply to this email."
        )
        body = f"{sal}\n\n{p1}\n\n{p2}\n\n{p3}\n\n{p4}\n\n{sig_en}"

    return betreff, body


def teaser_attached_in_body(body: str) -> bool:
    patterns = (
        r"\bAnhang\b",
        r"\bAttached\b",
        r"angehängt",
        r"im Anhang",
        r"teaser attached",
        r"füge.*Teaser.*bei",
        r"attach.*teaser",
        r"enclosed",
    )
    return any(re.search(p, body, re.I) for p in patterns)


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]
    lang_i = next(i for i, h in enumerate(headers) if h and "language" in str(h).lower())
    wrap = Alignment(wrap_text=True, vertical="top")

    cols = {h: i for i, h in enumerate(headers) if h}
    updated = 0

    for r in range(2, ws.max_row + 1):
        nr = ws.cell(r, cols["nr"] + 1).value
        if nr is None:
            continue
        name = ws.cell(r, cols["name"] + 1).value
        if name in (None, "k.A."):
            continue

        ansatz = ws.cell(r, cols["ansatz"] + 1).value
        if ansatz == "AUSSETZEN":
            continue

        nr = int(nr)
        meta = META.get(nr)
        if not meta:
            continue

        row = {
            "name": name,
            "contact_type": ws.cell(r, cols["contact_type"] + 1).value,
            "company": ws.cell(r, cols["company"] + 1).value,
        }
        de = language_is_de(ws.cell(r, lang_i + 1).value)
        betreff, body = build_email(row, meta, de, ansatz)

        ws.cell(r, cols["betreff"] + 1, betreff).alignment = wrap
        ws.cell(r, cols["email_body"] + 1, body).alignment = wrap
        anhang = "ja" if teaser_attached_in_body(body) else "nein"
        if "anhang" in cols:
            ws.cell(r, cols["anhang"] + 1, anhang)
        if "teaser_pfad" in cols and anhang == "nein":
            ws.cell(r, cols["teaser_pfad"] + 1, "")
        tpl = f"Ansatz_{ansatz}_v2_{'DE' if de else 'EN'}"
        if "template_ref" in cols:
            ws.cell(r, cols["template_ref"] + 1, tpl)
        updated += 1

    wb.save(XLSX)
    print(f"Updated {updated} email texts in {XLSX}")


if __name__ == "__main__":
    main()
