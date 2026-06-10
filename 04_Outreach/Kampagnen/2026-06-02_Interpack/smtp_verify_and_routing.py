#!/usr/bin/env python3
"""SMTP RCPT-TO check + routing text column for failed addresses."""
from __future__ import annotations

import importlib.util
import re
import smtplib
import socket
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
XLSX = BASE / "260605_CRM_Funnel_Interpack_List.xlsx"
SHEET = "Email_Outreach_Interpack_Warm"
ROUTING_COL = "email_routing_info@"
SMTP_COL = "smtp_check"
FROM_ADDR = "verify@loopforgelab.com"
TIMEOUT = 12
CACHE: dict[str, dict] = {}


def load_build():
    spec = importlib.util.spec_from_file_location(
        "build", BASE / "_build_email_outreach_sheet.py"
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


BUILD = load_build()


def mx_hosts(domain: str) -> list[str]:
    try:
        import dns.resolver  # type: ignore

        answers = dns.resolver.resolve(domain, "MX")
        pairs = sorted(
            (r.preference, str(r.exchange).rstrip(".")) for r in answers
        )
        return [h for _, h in pairs]
    except Exception:
        pass
    try:
        socket.getaddrinfo(domain, 25, proto=socket.IPPROTO_TCP)
        return [domain]
    except socket.gaierror:
        return []


def smtp_check(email: str) -> dict:
    email = email.strip().lower()
    if email in CACHE:
        return CACHE[email]
    if not email or email in ("k.a.", "—", "-") or "@" not in email:
        r = {"email": email, "status": "skip", "detail": "ungültig"}
        CACHE[email] = r
        return r

    domain = email.split("@", 1)[1]
    hosts = mx_hosts(domain)
    if not hosts:
        r = {"email": email, "status": "fail", "detail": "kein MX"}
        CACHE[email] = r
        return r

    last_err = ""
    for host in hosts[:3]:
        try:
            with smtplib.SMTP(host, 25, timeout=TIMEOUT) as smtp:
                smtp.ehlo_or_helo_if_needed()
                smtp.mail(FROM_ADDR)
                code, msg = smtp.rcpt(email)
                text = msg.decode(errors="replace") if isinstance(msg, bytes) else str(msg)
                if code in (250, 251):
                    r = {"email": email, "status": "ok", "detail": f"{host}: {code}"}
                    CACHE[email] = r
                    return r
                if code in (550, 551, 553, 554):
                    r = {"email": email, "status": "fail", "detail": f"{host}: {code} {text[:120]}"}
                    CACHE[email] = r
                    return r
                last_err = f"{host}: {code} {text[:120]}"
        except Exception as e:
            last_err = f"{host}: {type(e).__name__}: {e}"
            time.sleep(0.3)
            continue

    r = {"email": email, "status": "unknown", "detail": last_err or "keine Antwort"}
    CACHE[email] = r
    return r


def parse_alt_emails(raw) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[;,]\s*", str(raw))
    out = []
    for p in parts:
        p = p.strip()
        if not p or p.startswith("("):
            continue
        m = re.search(r"[\w.+-]+@[\w.-]+\.\w+", p)
        if m:
            out.append(m.group(0).lower())
    return out


def language_is_de(lang) -> bool:
    if not lang:
        return True
    return str(lang).strip().lower().startswith("deu")


def routing_text(row: dict, meta: dict, de: bool) -> str:
    """Generate info@ forwarding mail from existing email_body."""
    name = row.get("name", "")
    role = row.get("role", "")
    company = row.get("company", "")
    central = row.get("central_email") or "info@"
    body = row.get("email_body") or ""
    short = meta.get("short", company)

    if de:
        header = (
            f"Guten Tag,\n\n"
            f"könnten Sie diese Nachricht bitte an {name} ({role}) weiterleiten?\n\n"
            f"---\n\n"
        )
        footer = (
            f"\n\n---\n\n"
            f"Vielen Dank für die Weiterleitung.\n\n"
            f"Beste Grüße\nOlaf Pick\nLoopForgeLab GbR\nolaf@loopforgelab.com"
        )
        betreff_hint = f"Bitte weiterleiten an {name.split()[-1]} — Nach interpack ({short})"
    else:
        header = (
            f"Dear Sir or Madam,\n\n"
            f"could you please forward this message to {name} ({role})?\n\n"
            f"---\n\n"
        )
        footer = (
            f"\n\n---\n\n"
            f"Thank you for forwarding.\n\n"
            f"Best regards\nOlaf Pick\nLoopForgeLab GbR\nolaf@loopforgelab.com"
        )
        betreff_hint = f"Please forward to {name.split()[-1]} — follow-up after interpack ({short})"

    # Strip salutation/signoff from original if routing-style already
    core = body
    for cut in ("Beste Grüße", "Best regards", "Vielen Dank", "Thank you for forwarding"):
        if cut in core:
            core = core.split(cut)[0].strip()
    if core.startswith("Guten Tag") or core.startswith("Dear "):
        lines = core.split("\n", 2)
        core = lines[2].strip() if len(lines) > 2 else core

    return f"{header}{core}{footer}"


def smtp_status_label(primary: dict, alts: list[dict], row_email: str) -> str:
    if row_email and row_email.lower().startswith("info@"):
        tested = [a for a in alts if a["email"]]
        fails = [a for a in tested if a["status"] == "fail"]
        oks = [a for a in tested if a["status"] == "ok"]
        if oks:
            return f"routing — Direkt OK: {oks[0]['email']}"
        if fails:
            return f"routing — alle Direkt fail ({len(fails)} getestet)"
        return "routing — Direkt unbekannt (SMTP)"

    if primary["status"] == "ok":
        return "smtp OK"
    if primary["status"] == "fail":
        return f"smtp FAIL — {primary['detail'][:80]}"
    return f"smtp UNBEKANNT — {primary['detail'][:80]}"


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]

    def idx(name):
        for i, h in enumerate(headers):
            if h == name:
                return i
        return None

    col = {k: idx(k) for k in [
        "nr", "name", "company", "role", "contact_type", "email", "email_status",
        "central_email", "alt_email", "email_body", "betreff",
    ]}
    lang_col = next(i for i, h in enumerate(headers) if h and "language" in str(h).lower())

    # Add new columns if missing
    new_cols = [SMTP_COL, ROUTING_COL]
    for nc in new_cols:
        if nc not in headers:
            headers.append(nc)
            c = ws.cell(1, len(headers), nc)
            c.font = Font(bold=True)
            col[nc] = len(headers) - 1

    wrap = Alignment(wrap_text=True, vertical="top")
    results_log = []

    for r in range(2, ws.max_row + 1):
        nr_val = ws.cell(r, col["nr"] + 1).value
        if nr_val is None:
            continue
        nr = int(nr_val)
        name = ws.cell(r, col["name"] + 1).value
        if name in ("k.A.", None) or ws.cell(r, col["email"] + 1).value in (None, "k.A.", "—"):
            ws.cell(r, col[SMTP_COL] + 1, "aussetzen")
            ws.cell(r, col[ROUTING_COL] + 1, "—")
            continue

        row = {k: ws.cell(r, col[k] + 1).value for k in col if k in col and k not in new_cols}
        row["Language"] = ws.cell(r, lang_col + 1).value
        meta = BUILD.COMPANY_META.get(nr, {"short": row.get("company", "")})
        de = language_is_de(row["Language"])

        primary_email = (row.get("email") or "").strip().lower()
        alt_list = parse_alt_emails(row.get("alt_email"))

        # Test primary unless already info@ routing
        if primary_email.startswith("info@"):
            primary_res = {"email": primary_email, "status": "routing", "detail": "Ziel ist Zentrale"}
            # Test all direct candidates in alt_email
            to_test = alt_list
        else:
            to_test = [primary_email] + [a for a in alt_list if a != primary_email]

        tested = []
        for em in to_test:
            if em and em not in [t["email"] for t in tested]:
                tested.append(smtp_check(em))
                time.sleep(0.2)

        primary_res = next((t for t in tested if t["email"] == primary_email), tested[0] if tested else primary_res)
        alt_results = [t for t in tested if t["email"] != primary_email]

        status = smtp_status_label(primary_res, alt_results, primary_email)
        ws.cell(r, col[SMTP_COL] + 1, status)

        needs_routing = (
            primary_email.startswith("info@")
            or primary_res["status"] in ("fail", "unknown", "routing")
        )
        # If primary fails but an alt works, note it
        working_alt = next((a for a in alt_results if a["status"] == "ok"), None)

        if working_alt and not primary_email.startswith("info@"):
            ws.cell(r, col[SMTP_COL] + 1, f"smtp FAIL primär — nutze {working_alt['email']}")
            ws.cell(r, col["email"] + 1, working_alt["email"])
            ws.cell(r, col[ROUTING_COL] + 1, "")
            needs_routing = False

        if needs_routing:
            rt = routing_text(row, meta, de)
            cell = ws.cell(r, col[ROUTING_COL] + 1, rt)
            cell.alignment = wrap
            if primary_res["status"] == "fail":
                old = ws.cell(r, col["email_status"] + 1).value or ""
                ws.cell(r, col["email_status"] + 1, f"{old} | smtp FAIL".strip(" |"))
        else:
            ws.cell(r, col[ROUTING_COL] + 1, "")

        results_log.append({
            "row": r, "name": name, "primary": primary_email,
            "smtp_status": status, "tested": tested, "routing": needs_routing,
        })
        print(f"Row {r} {name}: {status}")

    # Column widths
    if ROUTING_COL in col:
        letter = get_column_letter(col[ROUTING_COL] + 1)
        ws.column_dimensions[letter].width = 72
    if SMTP_COL in col:
        letter = get_column_letter(col[SMTP_COL] + 1)
        ws.column_dimensions[letter].width = 36

    wb.save(XLSX)
    import json
    (BASE / "_smtp_results.json").write_text(
        json.dumps(results_log, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"Saved {XLSX}")


if __name__ == "__main__":
    main()
