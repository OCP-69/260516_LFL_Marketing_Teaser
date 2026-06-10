#!/usr/bin/env python3
"""Update Beumer rows in Email_Outreach_Interpack_Warm for info@ routing."""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

XLSX = Path(__file__).resolve().parent / "260605_CRM_Funnel_Interpack_List.xlsx"
SHEET = "Email_Outreach_Interpack_Warm"

SANDMARK = {
    "betreff": "Please forward to Mr Fredrik Sandmark — follow-up after interpack",
    "email_body": """Dear Sir or Madam,

could you please forward this message to Mr Fredrik Sandmark, CEO of BEUMER Maschinenfabrik GmbH & Co. KG?

My colleague Mouaz Al-Qudsi (CTO, LoopForgeLab) spoke with colleagues at your interpack stand. I am following up briefly to Mr Sandmark.

More quote variants with the same engineering capacity — Forge Engine turns BOM and CAD changes into reliable costs at design stage, without an ERP rollout.

What limits BEUMER Group more in 2026 — quote capacity or margin per order? If relevant: 15 minutes — https://calendly.com/olaf-pick/olaf-15min

Thank you for forwarding.

Best regards
Olaf Pick
LoopForgeLab GbR
olaf@loopforgelab.com""",
    "notizen": (
        "Direkt: fredrik.sandmark@ / f.sandmark@ blockiert (06/2026). "
        "CEO seit 03/2026 — Mailbox evtl. neu. "
        "Pattern-Kandidaten: Fredrik.Sandmark@beumergroup.com; fredrik.sandmark@beumer.com. "
        "Versand: routing via info@beumergroup.com"
    ),
    "alt_email": (
        "Fredrik.Sandmark@beumergroup.com; fredrik.sandmark@beumer.com; "
        "verena.breuer@beumergroup.com; jonas.jungmann@beumergroup.com"
    ),
}

KIRSCH = {
    "betreff": "Bitte weiterleiten an Herrn Christopher Kirsch — Nach interpack",
    "email_body": """Guten Tag,

könnten Sie diese Nachricht bitte an Herrn Christopher Kirsch, Divisional Director Global Innovation, weiterleiten?

Mein Kollege Mouaz Al-Qudsi (CTO, LoopForgeLab) war auf der interpack mit Kollegen von Ihrem Stand im Gespräch. Herr Kirsch ist auf Ihrer Seite zur digitalen Transformation als Ansprechpartner genannt — daher diese Ebene.

Weniger Iterationsschleifen vor Freigabe: Varianten aus BOM, CAD und Einkauf sofort in Euro — kein neues ERP.

Gehören teure Änderungen nach Freigabe bei Ihnen zum Alltag — oder eher zur Ausnahme? 20 Min. reichen für einen ersten Abgleich: https://calendly.com/olaf-pick/olaf-15min

Vielen Dank für die Weiterleitung.

Beste Grüße
Olaf Pick
LoopForgeLab GbR
olaf@loopforgelab.com""",
    "notizen": (
        "Direkt: christopher.kirsch@ / c.kirsch@ nicht zustellbar. "
        "Offiziell: beumergroup.com/about-us/digital-transformation (Name, kein Mail). "
        "Pattern-Kandidaten: Christopher.Kirsch@beumergroup.com; christopher.kirsch@beumer.com. "
        "Versand: routing via info@beumergroup.com"
    ),
    "alt_email": (
        "Christopher.Kirsch@beumergroup.com; christopher.kirsch@beumer.com; "
        "verena.breuer@beumergroup.com"
    ),
}


def col_idx(headers, name):
    for i, h in enumerate(headers):
        if h == name:
            return i
    return None


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]
    wrap = Alignment(wrap_text=True, vertical="top")

    idx = {
        "name": col_idx(headers, "name"),
        "email": col_idx(headers, "email"),
        "status": col_idx(headers, "email_status"),
        "betreff": col_idx(headers, "betreff"),
        "body": col_idx(headers, "email_body"),
        "alt": col_idx(headers, "alt_email"),
        "notizen": col_idx(headers, "notizen"),
        "versand": col_idx(headers, "versand_status"),
        "template": col_idx(headers, "template_ref"),
    }

    updated = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, idx["name"] + 1).value
        if name == "Fredrik Sandmark":
            data = SANDMARK
        elif name == "Christopher Kirsch":
            data = KIRSCH
        else:
            continue

        ws.cell(r, idx["email"] + 1, "info@beumergroup.com")
        ws.cell(r, idx["status"] + 1, "routing (info@) — Direktmail blockiert")
        ws.cell(r, idx["betreff"] + 1, data["betreff"]).alignment = wrap
        ws.cell(r, idx["body"] + 1, data["email_body"]).alignment = wrap
        ws.cell(r, idx["alt"] + 1, data["alt_email"])
        ws.cell(r, idx["notizen"] + 1, data["notizen"]).alignment = wrap
        ws.cell(r, idx["versand"] + 1, "routing vorbereitet")
        ws.cell(r, idx["template"] + 1, "Routing_info@_Beumer")
        updated.append(name)

    wb.save(XLSX)
    print(f"Updated: {updated}")


if __name__ == "__main__":
    main()
