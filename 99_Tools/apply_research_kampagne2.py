#!/usr/bin/env python3
"""Apply firm-by-firm research to Kontakte_Email (Kampagne 2) + SMTP check."""
from __future__ import annotations

import csv
import json
import re
import smtplib
import socket
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "260615_Kampagne_2.xlsx"
OUT_DIR = ROOT / "04_Outreach/Kampagnen/2026-06-15_Kampagne_2"
CSV_OUT = OUT_DIR / "260615_Kontakte_Email.csv"
SHEET = "Kontakte_Email"
FROM_ADDR = "verify@loopforgelab.com"
TIMEOUT = 12
SMTP_CACHE: dict[str, dict] = {}

HEADERS = [
    "nr", "segment", "company", "location", "website", "contact_type", "name", "role",
    "email", "email_status", "email_pattern", "smtp_check", "source_url", "source_type",
    "alt_email", "central_email", "central_phone", "linkedin_url", "produkte_kurz",
    "zielkunden_kurz", "branche_kontext", "website_fakt", "hook_fakt", "hook_quelle",
    "hook_lens", "pain_kundensprache", "email_betreff", "email_body", "status", "remarks",
]

# Key: (nr, contact_type) -> field overrides
RESEARCH: dict[tuple[int, str], dict] = {
    # --- 1 GEBHARDT ---
    (1, "CEO/GF"): {
        "website": "https://www.gebhardt-group.com",
        "central_email": "info@gebhardt-group.com",
        "email_pattern": "{First}.{Last}@gebhardt-group.com",
        "name": "Marco Gebhardt",
        "role": "Geschäftsführer / Inhaber (3. Generation)",
        "email": "marco.gebhardt@gebhardt-group.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://gebhardt-group.com/en/company/gebhardt-group.html",
        "source_type": "firmenwebsite",
        "alt_email": "info@gebhardt-group.com",
        "website_fakt": "LogiMAT 2026; US-Produktionsausbau Streetsboro (Okt 2025); StoreBiter-Shuttle/AMR-Portfolio.",
        "hook_fakt": "Mit dem US-Manufacturing-Ausbau und LogiMAT 2026 skalieren Sie parallel mehrere kundenspezifische Layout-Varianten — typisch für ETO-Intralogistik.",
        "hook_quelle": "https://us.gebhardt-group.com/company/news/",
        "email_betreff": "GEBHARDT — Angebotskapazität oder Marge 2026?",
        "status": "recherche_fertig",
        "remarks": "GF laut Presse/Management; Pattern aus Impressum-Kollegen prüfen",
    },
    (1, "Engineering"): {
        "website": "https://www.gebhardt-group.com",
        "central_email": "info@gebhardt-group.com",
        "email_pattern": "{First}.{Last}@gebhardt-group.com",
        "name": "Leitung Konstruktion & Entwicklung",
        "role": "Engineering (kein öffentlicher Namens-Kontakt)",
        "email": "info@gebhardt-group.com",
        "email_status": "routing — kein Direktkontakt",
        "source_url": "https://jobs.gebhardt-group.com/",
        "source_type": "stellenausschreibung",
        "alt_email": "marco.gebhardt@gebhardt-group.com",
        "website_fakt": "Neuer Geschäftsbereich Retrofit (Director-Stelle); Controls/Robotics-Teams in Sinsheim.",
        "hook_fakt": "Retrofit und Warehouse-Robotics erweitern Ihr Engineering — parallel laufen oft Layout- und Steuerungsvarianten in derselben RFQ-Phase.",
        "hook_quelle": "https://jobs.gebhardt-group.com/",
        "email_betreff": "GEBHARDT — Varianten vor Freigabe?",
        "status": "routing_erforderlich",
        "remarks": "Kein öffentlicher Leiter Konstruktion; Weiterleitung via info@",
    },
    # --- 2 AUMUND ---
    (2, "CEO/GF"): {
        "name": "Dr.-Ing. Pietro de Michieli",
        "role": "CEO / Geschäftsführer",
        "email": "pietro.demichieli@aumund.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.aumund.com/en/company/management/",
        "source_type": "firmenwebsite",
        "alt_email": "info@aumund.com",
        "website_fakt": "Nischenführer Schwerlastförderung; Großprojekte Zement/Bergbau/Häfen weltweit.",
        "hook_fakt": "Bei Großprojekt-ETO mit langen Angebotszyklen frage ich mich, ob Layout- oder Detailvarianten Ihre Senior-Ingenieure stärker binden.",
        "hook_quelle": "https://www.aumund.com",
        "email_betreff": "AUMUND — Engineering-Zeit in der Kalkulation?",
        "status": "recherche_fertig",
    },
    (2, "Engineering"): {
        "name": "Reiner Furthmann",
        "role": "CTO",
        "email": "reiner.furthmann@aumund.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.aumund.com/en/company/management/",
        "source_type": "firmenwebsite",
        "alt_email": "info@aumund.com",
        "website_fakt": "Schwerlast- und Bulk-Material-Handling; variantenreiche Anlagenkonfigurationen pro Projekt.",
        "hook_fakt": "Bei Schwerlast-ETO-Projekten höre ich oft: Die teuersten Überraschungen entstehen vor der Freigabe — wenn Layout- und Detailvarianten nicht in derselben Kostensicht laufen.",
        "hook_quelle": "https://www.aumund.com/en/company/management/",
        "email_betreff": "AUMUND — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 3 STIWA ---
    (3, "CEO/GF"): {
        "name": "Michael Fuchshuber",
        "role": "CEO STIWA Automation",
        "email": "michael.fuchshuber@stiwa.com",
        "email_status": "verifiziert (Website)",
        "source_url": "https://www.stiwa.com/en/manufacturing/contact",
        "source_type": "firmenwebsite",
        "alt_email": "office@stiwa.com",
        "website_fakt": "FAST-Montagezellen, High-Performance-Automation; Investitionen in KI-gestützte Montage.",
        "hook_fakt": "Jede Sonderzelle ist ein Unikat — ich frage mich, ob frühere Projektdaten bei neuen RFQs wirklich in der Kalkulation ankommen.",
        "hook_quelle": "https://www.stiwa.com",
        "email_betreff": "STIWA — Angebotskapazität oder Marge?",
        "status": "recherche_fertig",
    },
    (3, "Engineering"): {
        "name": "Gerhard Berer",
        "role": "Bereichsleiter Engineering",
        "email": "gerhard.berer@stiwa.com",
        "email_status": "abgeleitet (hoch)",
        "source_url": "https://www.linkedin.com/in/gerhard-berer-656370150",
        "source_type": "linkedin",
        "alt_email": "office@stiwa.com",
        "website_fakt": "Engineering für Montageautomation und Sonderzellen; dezentrale Bereichsleitungen.",
        "hook_fakt": "Bei kundenspezifischen FAST-Zellen: Wo verlieren Sie mehr Zeit — Varianten durchrechnen oder auf Einkauf/Kalkulation warten?",
        "hook_quelle": "https://www.linkedin.com/in/gerhard-berer-656370150",
        "email_betreff": "STIWA — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 4 DS Automotion ---
    (4, "CEO/GF"): {
        "name": "Wolfgang Hillinger",
        "role": "CEO",
        "email": "wolfgang.hillinger@ds-automotion.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.ds-automotion.com/en/company/about-us/",
        "source_type": "firmenwebsite",
        "alt_email": "office@ds-automotion.com",
        "website_fakt": "SSI-Schäfer-Tochter seit 2023; Schwerlast-AGV und Flottensteuerung aus Linz.",
        "hook_fakt": "Kundenspezifische AGV-Flotten mit variantenreicher Konfiguration — bindet Engineering in der Angebotsphase Kapazität oder Marge?",
        "hook_quelle": "https://www.ds-automotion.com",
        "email_betreff": "DS Automotion — RFQ-Kapazität 2026?",
        "status": "recherche_fertig",
    },
    (4, "Engineering"): {
        "name": "Thomas Gschwentenwein",
        "role": "CTO / Bereichsleiter Technologie & Entwicklung",
        "email": "thomas.gschwentenwein@ds-automotion.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.ds-automotion.com/en/company/about-us/",
        "source_type": "firmenwebsite",
        "alt_email": "office@ds-automotion.com",
        "website_fakt": "Schwerlast-AGV-Entwicklung; kundenspezifische Fahrzeug- und Flottenvarianten.",
        "hook_fakt": "Bei AGV-ETO: Trifft „Nacharbeit nach Freigabe“ bei Ihnen zu — oder stockt eher die Durchlaufzeit der Varianten in der Konstruktion?",
        "hook_quelle": "https://www.ds-automotion.com",
        "email_betreff": "DS Automotion — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 5 Kolbus ---
    (5, "CEO/GF"): {
        "name": "Michael R. Bach",
        "role": "CEO (seit Jan 2026)",
        "email": "michael.bach@kolbus.de",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://kolbus.com/de/unternehmen/news/news/new-management-team-strengthens-kolbus-group-gmbh-in-rahden",
        "source_type": "presse",
        "alt_email": "info@kolbus.de",
        "website_fakt": "Neues GL-Team seit Jan 2026 (Bach, Galburt, Sieckmann, Kolwey); Max Valier Holding; Fokus Wellpappe/Print.",
        "hook_fakt": "Mit neuem CEO und frischem GL-Team nach der Restrukturierung: Ist Kalkulationswissen institutionalisiert — oder noch stark in Köpfen gebunden?",
        "hook_quelle": "https://kolbus.com/de/unternehmen/news/",
        "email_betreff": "Kolbus — was limitiert Wachstum 2026?",
        "status": "recherche_fertig",
    },
    (5, "Engineering"): {
        "name": "Mathias Sieckmann",
        "role": "CCO & CTO (Geschäftsleitung)",
        "email": "mathias.sieckmann@kolbus.de",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.thepackagingportal.com/industry-news/new-senior-management-team-at-kolbus/",
        "source_type": "presse",
        "alt_email": "info@kolbus.de",
        "website_fakt": "Case-Maker und Buchbinderei; variantenreiche Maschinenkonfigurationen für Converter.",
        "hook_fakt": "Bei Wellpappen-/Print-Maschinen: Werden Format- und Materialvarianten bei Ihnen eher vor oder nach der Freigabe in derselben Kostensicht bewertet?",
        "hook_quelle": "https://kolbus.com",
        "email_betreff": "Kolbus — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 6 W&H ---
    (6, "CEO/GF"): {
        "name": "Dr. Falco Paepenmüller",
        "role": "CEO (vormals CTO)",
        "email": "falco.paepenmueller@wuh-group.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.wh.group/de/unternehmen/w_h_group/",
        "source_type": "firmenwebsite",
        "linkedin_url": "",
        "alt_email": "info@wuh-group.com",
        "website_fakt": "Paper-based flexible packaging (PPWR-Treiber); Technology Center Lengerich; ~3.400 MA.",
        "hook_fakt": "Materialumstellung Folie→Papier treibt neue Maschinenvarianten — rutschen FMCG-Specs und PPWR-Themen früher in Ihre Angebotsphase?",
        "hook_quelle": "https://www.wh.group/de/unternehmen/w_h_group/",
        "email_betreff": "W&H — Angebotskapazität oder Marge?",
        "status": "recherche_fertig",
        "remarks": "Korrigiert: Paepenmüller CEO seit 2024 (nicht Nienkemper)",
    },
    (6, "Engineering"): {
        "name": "Dr. Björn Feldhaus",
        "role": "Director R&D & Sustaining Engineering",
        "email": "bjoern.feldhaus@wuh-group.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.linkedin.com/in/dr-bj%C3%B6rn-feldhaus-4b4327153",
        "source_type": "linkedin",
        "alt_email": "info@wuh-group.com",
        "website_fakt": "R&D für Flexodruck/Extrusion; technotrans-Partnerschaft Farbversorgung Flexodruck.",
        "hook_fakt": "Bei Flexo-/Extrusions-Varianten: Ist das bei Ihnen eher Ausnahme — oder Alltag, dass jede Sonderanfrage fast bei Null startet trotz vorhandener CAD-Daten?",
        "hook_quelle": "https://www.wirtschaft-aktuell.de/news/technotrans-und-wh-vertiefen-partnerschaft",
        "email_betreff": "W&H — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 7 Theegarten ---
    (7, "CEO/GF"): {
        "name": "Thomas Rother",
        "role": "Geschäftsführer",
        "email": "thomas.rother@theegarten-pactec.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.theegarten-pactec.com",
        "source_type": "firmenwebsite",
        "alt_email": "info@theegarten-pactec.com",
        "website_fakt": "Neubau Produktionshalle Dresden; F&E-Digitalisierung; Export Süßwaren-OEMs.",
        "hook_fakt": "Nach Kapazitätsausbau Dresden laufen parallele Süßwaren-RFQs — bindet Variantenkalkulation Senior-Ingenieure?",
        "hook_quelle": "https://www.theegarten-pactec.com",
        "email_betreff": "Theegarten-Pactec — RFQ-Welle 2026?",
        "status": "recherche_fertig",
    },
    (7, "Engineering"): {
        "name": "Benedikt Schulte",
        "role": "Director Engineering & Development",
        "email": "benedikt.schulte@theegarten-pactec.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.theegarten-pactec.com",
        "source_type": "firmenwebsite",
        "alt_email": "info@theegarten-pactec.com",
        "website_fakt": "High-Speed Wrapper für Ferrero, Lindt, Mars; Servo-Plattformen.",
        "hook_fakt": "Formatvarianten bei High-Speed-Linien: Wo verlieren Sie mehr Zeit — Varianten durchrechnen oder auf Kalkulation warten?",
        "hook_quelle": "https://www.theegarten-pactec.com",
        "email_betreff": "Theegarten — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 8 Optima ---
    (8, "CEO/GF"): {
        "name": "Dr. Stefan König",
        "role": "CEO Optima Packaging Group",
        "email": "stefan.koenig@optima-packaging.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.optima-packaging.com",
        "source_type": "firmenwebsite",
        "alt_email": "info@optima-packaging.com",
        "website_fakt": "Führungswechsel; US-Expansion; neues Trainingszentrum Schwäbisch Hall; INOVA-Portfolio.",
        "hook_fakt": "GMP-Varianten und PPWR-Anforderungen rutschen bei Pharma-/Consumer-Linien oft früher in die Angebotsphase — institutionalisiert oder in Köpfen?",
        "hook_quelle": "https://www.optima-packaging.com",
        "email_betreff": "Optima — Angebotskapazität oder Marge?",
        "status": "recherche_fertig",
        "remarks": "Korrigiert: König CEO (Bühler Gesellschafter, nicht Primärkontakt)",
    },
    (8, "Engineering"): {
        "name": "Tobias Meinikheim",
        "role": "Director of Engineering",
        "email": "tobias.meinikheim@optima-packaging.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.optima-packaging.com",
        "source_type": "firmenwebsite",
        "alt_email": "info@optima-packaging.com",
        "website_fakt": "Pharma/Cosmetics/Consumer-Linien; Engineering Change Orders und Digital Twin.",
        "hook_fakt": "Bei GMP-Linien: Trifft „Nacharbeit nach Freigabe“ zu — oder stockt die Durchlaufzeit der Varianten in der Konstruktion?",
        "hook_quelle": "https://www.optima-packaging.com",
        "email_betreff": "Optima — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 9 Schubert (bereits recherchiert) ---
    (9, "CEO/GF"): {
        "name": "Ralf Schubert",
        "role": "Geschäftsführer (Technologie-Division)",
        "email": "r.schubert@gerhard-schubert.de",
        "email_status": "abgeleitet (hoch)",
        "source_url": "https://www.schubert.group/en/imprint/",
        "source_type": "impressum",
        "alt_email": "info@schubert.group",
        "website_fakt": "Schubert-Pharma Expansion; TLM-Modularsystem; AI Vision.",
        "hook_fakt": "Modulare TLM-Linien erzeugen hohe Variantenvielfalt — limitiert RFQ-Tempo Ihre Engineering-Kapazität oder die Marge pro Auftrag?",
        "hook_quelle": "https://www.schubert.group",
        "email_betreff": "Schubert — RFQ-Tempo vs. Engineering?",
        "status": "recherche_fertig",
    },
    (9, "Engineering"): {
        "name": "Marcus Schindler",
        "role": "Bereichsleiter Konstruktion",
        "email": "m.schindler@gerhard-schubert.de",
        "email_status": "abgeleitet (hoch)",
        "source_url": "https://www.schubert.group/en/group/locations/",
        "source_type": "firmenwebsite",
        "linkedin_url": "https://www.linkedin.com/in/marcus-schindler-789554131",
        "alt_email": "t.neumann@gerhard-schubert.de",
        "website_fakt": "Pick-and-Place TLM; variantenreiche Food-/Pharma-Linien.",
        "hook_fakt": "Bei modularen Linien: Werden Material- und Formatvarianten vor der Freigabe in derselben Kostensicht bewertet?",
        "hook_quelle": "https://www.schubert.group",
        "email_betreff": "Schubert — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 10 Bielomatik ---
    (10, "CEO/GF"): {
        "company": "Bielomatik GmbH",
        "location": "DE, Römerstein",
        "name": "Marco Imperatore",
        "role": "Geschäftsführer",
        "email": "marco.imperatore@bielomatik.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.bielomatik.com",
        "source_type": "firmenwebsite",
        "alt_email": "info@bielomatik.com",
        "website_fakt": "Neuffen KG i.L. abgewickelt; operative Gesellschaft Römerstein (Paper Converting, Friction Welding).",
        "hook_fakt": "Nach Rechtsform-/Standortwechsel: Ist Kalkulationswissen für Paper-Converting-Varianten institutionalisiert?",
        "hook_quelle": "https://www.bielomatik.com",
        "email_betreff": "Bielomatik — Angebotskapazität 2026?",
        "status": "recherche_fertig",
        "remarks": "Korrigiert: Römerstein/Imperatore (nicht Neuffen KG)",
    },
    (10, "Engineering"): {
        "company": "Bielomatik GmbH",
        "location": "DE, Römerstein",
        "name": "Stefan Carstensen",
        "role": "Leiter Entwicklung / Konstruktion (Paper Converting)",
        "email": "stefan.carstensen@bielomatik.com",
        "email_status": "abgeleitet (niedrig)",
        "source_url": "https://www.bielomatik.com",
        "source_type": "firmenwebsite",
        "alt_email": "info@bielomatik.com",
        "website_fakt": "Paper Converting und Tissue-Linien; ETO-Varianten pro Kunde.",
        "hook_fakt": "Bei Paper-Converting-Varianten: Frühe Kostensicht im Design — oder Nacharbeit nach Freigabe?",
        "hook_quelle": "https://www.bielomatik.com",
        "email_betreff": "Bielomatik — Varianten vor Freigabe?",
        "status": "recherche_fertig",
        "remarks": "Carstensen aus älteren Quellen; Namen vor Versand verifizieren",
    },
    # --- 11 EREMA ---
    (11, "CEO/GF"): {
        "name": "Manfred Hackl",
        "role": "CEO",
        "email": "manfred.hackl@erema.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.erema.com",
        "source_type": "firmenwebsite",
        "alt_email": "public.relations@erema-group.com",
        "website_fakt": "Marktführer Plastic Recycling; PPWR und PCR-Anteile als Kundenspecs.",
        "hook_fakt": "PCR-Anteil als Kundenspec treibt Maschinenvarianten — rutschen diese früher in Ihre Angebotsphase als früher?",
        "hook_quelle": "https://www.erema.com",
        "email_betreff": "EREMA — RFQ-Kapazität oder Marge?",
        "status": "recherche_fertig",
    },
    (11, "Engineering"): {
        "name": "Dr. Horst Wolfsgruber",
        "role": "Leiter Entwicklung / CTO",
        "email": "horst.wolfsgruber@erema.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.erema.com",
        "source_type": "firmenwebsite",
        "alt_email": "office@erema.at",
        "website_fakt": "INTAREMA/VACUREMA; variantenreiche Recycling-Linien nach Materialmix.",
        "hook_fakt": "Bei PCR-Recycling-Linien: Werden Materialmix-Varianten vor der Freigabe in derselben Kostensicht bewertet?",
        "hook_quelle": "https://www.erema.com",
        "email_betreff": "EREMA — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 12 Bobst ---
    (12, "CEO/GF"): {
        "name": "Jean-Pascal Bobst",
        "role": "CEO Bobst Group",
        "email": "gudrun.alex@bobst.com",
        "email_status": "routing — Presse",
        "source_url": "https://www.bobst.com",
        "source_type": "presse",
        "alt_email": "leonard.badet@bobst.com",
        "website_fakt": "Konzern Folding Carton/Corrugated/Flexible; Sustainability & Digital Services.",
        "hook_fakt": "Bei Konfigurationsvielfalt über Divisionen: Bottleneck Angebot oder Engineering-Kapazität?",
        "hook_quelle": "https://www.bobst.com",
        "email_betreff": "Bobst — Angebotskapazität 2026?",
        "status": "routing_erforderlich",
        "remarks": "Lead B · Konzern · Presse-Routing (05_Konzern_ueber_Presse.md)",
    },
    (12, "Engineering"): {
        "name": "Leonard Badet",
        "role": "CTO",
        "email": "leonard.badet@bobst.com",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.bobst.com",
        "source_type": "firmenwebsite",
        "alt_email": "gudrun.alex@bobst.com",
        "website_fakt": "Globaler Maschinenbau-Konzern; hohe Konfigurationsvielfalt.",
        "hook_fakt": "Konfigurationsvarianten in Folding Carton/Flexible: Nacharbeit nach Freigabe oder Varianten-Durchlaufzeit?",
        "hook_quelle": "https://www.bobst.com",
        "email_betreff": "Bobst — Varianten vor Freigabe?",
        "status": "recherche_fertig",
    },
    # --- 13 Senning ---
    (13, "CEO/GF"): {
        "company": "Senning GmbH",
        "name": "Annette Bengs",
        "role": "Geschäftsführerin (3. Generation)",
        "email": "annette.bengs@senning.de",
        "email_status": "abgeleitet (mittel)",
        "source_url": "https://www.optima-packaging.com/de/newsroom/presse/pressemitteilungen/senning-wird-teil-der-optima-unternehmensgruppe",
        "source_type": "presse",
        "alt_email": "info@senning.de",
        "website_fakt": "Teil der Optima-Gruppe seit 2020; ~50 MA; 85% Export Tissue/Paper.",
        "hook_fakt": "Als familiengeführtes ETO mit ~50 MA: Bindet Kalkulation am GF/Ingenieur — oder ist das institutionalisiert?",
        "hook_quelle": "https://www.senning.de",
        "email_betreff": "Senning — Angebotskapazität?",
        "status": "recherche_fertig",
        "remarks": "Optima-Konzern; nicht Hans Bühler kontaktieren (Optima-CEO)",
    },
    (13, "Engineering"): {
        "company": "Senning GmbH",
        "name": "Oliver Rebstock",
        "role": "Geschäftsführer / Technik & Operations",
        "email": "oliver.rebstock@senning.de",
        "email_status": "abgeleitet (niedrig)",
        "source_url": "https://www.senning.de",
        "source_type": "firmenwebsite",
        "alt_email": "info@senning.de",
        "website_fakt": "Tissue-/Paper-Verpackungsmaschinen; ETO pro Kunde.",
        "hook_fakt": "Bei Tissue-Linien: Varianten vor Freigabe in belastbare Kosten — oder Nacharbeit?",
        "hook_quelle": "https://www.senning.de",
        "email_betreff": "Senning — Varianten vor Freigabe?",
        "status": "recherche_fertig",
        "remarks": "Kleines ETO; Rebstock als technischer GF-Kontakt",
    },
    # --- 14 Fill ---
    (14, "CEO/GF"): {
        "company": "Fill GmbH",
        "name": "Andreas Fill",
        "role": "CEO",
        "email": "andreas.fill@fill.co.at",
        "email_status": "abgeleitet (mittel)",
        "email_pattern": "{first}.{last}@fill.co.at",
        "source_url": "https://www.fill.co.at",
        "source_type": "firmenwebsite",
        "alt_email": "info@fill.co.at",
        "website_fakt": "Sondermaschinenbau Metall/Kunststoff/Holz (nicht klassischer Verpackungs-OEM); Hexagon-Humanoid-Partnerschaft 2026.",
        "hook_fakt": "Bei kundenspezifischen Sondermaschinen: Wo stocken RFQ-Schleifen — Variantenkalkulation oder Freigabe?",
        "hook_quelle": "https://www.fill.co.at",
        "email_betreff": "Fill — RFQ-Durchlaufzeit?",
        "status": "recherche_fertig",
        "remarks": "Kein Packaging-OEM; Segment (E) stärker als (V); fillpack.de separat",
    },
    (14, "Engineering"): {
        "company": "Fill GmbH",
        "name": "Alois Wiesinger",
        "role": "CTO",
        "email": "alois.wiesinger@fill.co.at",
        "email_status": "abgeleitet (mittel)",
        "email_pattern": "{first}.{last}@fill.co.at",
        "source_url": "https://www.fill.co.at",
        "source_type": "firmenwebsite",
        "alt_email": "info@fill.co.at",
        "website_fakt": "ETO-Sondermaschinen; hohe Variantenkomplexität pro Projekt.",
        "hook_fakt": "Bei Sondermaschinen-ETO: Trifft „Nacharbeit nach Freigabe“ zu — oder Varianten-Durchlaufzeit?",
        "hook_quelle": "https://www.fill.co.at",
        "email_betreff": "Fill — Varianten vor Freigabe?",
        "status": "recherche_fertig",
        "remarks": "Sondermaschinenbau, nicht Verpackungsmaschinenbau",
    },
}


def last_name(full_name: str) -> str:
    clean = re.sub(r"^(Dr\.|Dr\.-Ing\.|Prof\.)\s*", "", full_name.strip())
    parts = [p for p in clean.split() if p]
    return parts[-1] if parts else full_name


def is_female(name: str) -> bool:
    first = name.split()[0].lower() if name else ""
    return first in {"annette", "cristina", "sabine", "petra", "andrea"}


def salutation(name: str) -> str:
    if not name or name.startswith("Leitung"):
        return "Guten Tag,"
    ln = last_name(name)
    if is_female(name):
        return f"Guten Tag Frau {ln},"
    return f"Guten Tag Herr {ln},"


def company_short(company: str) -> str:
    c = company.split("(")[0].strip()
    for cut in (" GmbH", " KG", " SA", " Group", " & Co."):
        if cut in c:
            c = c.split(cut)[0].strip()
    return c.split()[0] if c else company


def build_ceo_body(row: dict) -> str:
    company = row["company"]
    short = company_short(company)
    seg = row.get("segment", "")
    hook = row.get("hook_fakt", "")
    intro = (
        "ich schreibe Ihnen, weil wir bei LoopForgeLab mit wenigen ETO-Herstellern "
        "im Verpackungsmaschinenbau den Austausch zu Kalkulation und Varianten im Design "
        "suchen — ohne IT-Großprojekt."
        if seg == "Packaging" and "Fill" not in company
        else "ich schreibe Ihnen, weil wir bei LoopForgeLab mit wenigen ETO-Herstellern "
        "in der Intralogistik und im Sondermaschinenbau den Austausch zu Kalkulation "
        "und Varianten im Design suchen — ohne IT-Großprojekt."
    )
    return f"""{salutation(row.get("name", ""))}

{intro}

{hook}

Aus Gesprächen mit anderen Geschäftsführern im Sondermaschinenbau höre ich oft eine Spannung — ich würde gern prüfen, ob das bei Ihnen überhaupt zutrifft:

Entweder limitiert **Angebotskapazität** das Wachstum (zu viel Engineering in der Kalkulation), oder die **Marge** pro gewonnenem Auftrag (Puffer, Nachkalkulation, zu späte Kostensicht).

**Was ist bei {short} 2026 der schwerere Engpass — oder liegt Ihr Fokus ganz woanders?**

Eine kurze Einordnung reicht mir — ich möchte nicht ins Leere schreiben.

Beste Grüße
Olaf Pick
LoopForgeLab GbR
olaf@loopforgelab.com"""


def build_eng_body(row: dict) -> str:
    company = row["company"]
    short = company_short(company)
    hook = row.get("hook_fakt", "")
    return f"""{salutation(row.get("name", ""))}

ich gehe gezielt auf Engineering-Leitungen im Sondermaschinenbau zu — Hintergrund: wie Varianten und Kosten **vor der Freigabe** zusammenlaufen.

{hook}

In anderen ETO-Betrieben höre ich häufig: Die teuersten Überraschungen kommen nicht in der Montage selbst, sondern **vor der Freigabe** — wenn Varianten und Materialwechsel noch nicht mit derselben Kostensicht bewertet werden.

**Trifft „Nacharbeit / Änderungsaufträge nach Freigabe“ bei {short} eher zu — oder ist der Engpass eher die Durchlaufzeit der Varianten in der Konstruktion?**

Ich bin an Ihrer Einordnung interessiert, nicht an einem Produkt-Pitch.

Beste Grüße
Olaf Pick
LoopForgeLab GbR
olaf@loopforgelab.com"""


def routing_body(row: dict) -> str:
    name = row.get("name", "")
    role = row.get("role", "")
    core = build_ceo_body(row) if row.get("contact_type") == "CEO/GF" else build_eng_body(row)
  # strip salutation for routing wrapper
    lines = core.split("\n", 2)
    inner = lines[2] if len(lines) > 2 else core
    for cut in ("Beste Grüße", "Best regards"):
        if cut in inner:
            inner = inner.split(cut)[0].strip()
    target = name if not name.startswith("Leitung") else f"die {role}"
    return f"""Guten Tag,

könnten Sie diese Nachricht bitte an {target} weiterleiten?

---

{inner}

---

Vielen Dank für die Weiterleitung.

Beste Grüße
Olaf Pick
LoopForgeLab GbR
olaf@loopforgelab.com"""


def mx_hosts(domain: str) -> list[str]:
    try:
        import dns.resolver  # type: ignore

        answers = dns.resolver.resolve(domain, "MX")
        pairs = sorted((r.preference, str(r.exchange).rstrip(".")) for r in answers)
        return [h for _, h in pairs]
    except Exception:
        pass
    try:
        socket.getaddrinfo(domain, 25, proto=socket.IPPROTO_TCP)
        return [domain]
    except socket.gaierror:
        return []


def smtp_check(email: str) -> dict:
    email = (email or "").strip().lower()
    if email in SMTP_CACHE:
        return SMTP_CACHE[email]
    if not email or "@" not in email:
        r = {"email": email, "status": "skip", "detail": "ungültig"}
        SMTP_CACHE[email] = r
        return r
    domain = email.split("@", 1)[1]
    hosts = mx_hosts(domain)
    if not hosts:
        r = {"email": email, "status": "fail", "detail": "kein MX"}
        SMTP_CACHE[email] = r
        return r
    last_err = ""
    for host in hosts[:3]:
        try:
            with smtplib.SMTP(host, 25, timeout=TIMEOUT) as smtp:
                smtp.ehlo_or_helo_if_needed()
                smtp.mail(FROM_ADDR)
                code, msg = smtp.rcpt(email)
                text = msg.decode(errors="replace") if isinstance(msg, bytes) else str(msg)
                if code in (250, 251):
                    r = {"email": email, "status": "ok", "detail": f"{host}: {code}"}
                    SMTP_CACHE[email] = r
                    return r
                if code in (550, 551, 553, 554):
                    r = {"email": email, "status": "fail", "detail": f"{host}: {code} {text[:100]}"}
                    SMTP_CACHE[email] = r
                    return r
                last_err = f"{host}: {code} {text[:100]}"
        except Exception as e:
            last_err = f"{host}: {type(e).__name__}"
            time.sleep(0.25)
    r = {"email": email, "status": "unknown", "detail": last_err or "keine Antwort"}
    SMTP_CACHE[email] = r
    return r


def parse_alt_emails(raw) -> list[str]:
    if not raw:
        return []
    out = []
    for p in re.split(r"[;,]\s*", str(raw)):
        m = re.search(r"[\w.+-]+@[\w.-]+\.\w+", p)
        if m:
            out.append(m.group(0).lower())
    return out


def is_ip_block(detail: str) -> bool:
    d = (detail or "").lower()
    return any(
        x in d
        for x in (
            "spamhaus",
            "barracuda",
            "client host",
            "blocked using",
            "service unavailable; client",
        )
    )


def smtp_label(primary: str, results: list[dict], is_routing: bool) -> str:
    if is_routing:
        alts = [r for r in results if r["email"] != primary.lower()]
        oks = [r for r in alts if r["status"] == "ok"]
        fails = [r for r in alts if r["status"] == "fail"]
        blocked = [r for r in fails if is_ip_block(r.get("detail", ""))]
        if oks:
            return f"routing — Direkt OK: {oks[0]['email']}"
        if blocked and len(blocked) == len(fails):
            return "routing — SMTP blockiert (IP) — manuell prüfen"
        if fails and not oks:
            return f"routing — Direkt fail ({len(fails)} getestet)"
        return "routing — Direkt unbekannt"
    pr = next((r for r in results if r["email"] == primary.lower()), None)
    if not pr:
        return "nicht getestet"
    if pr["status"] == "ok":
        return "smtp OK"
    if pr["status"] == "fail":
        if is_ip_block(pr.get("detail", "")):
            return "smtp BLOCKIERT (IP/Spamhaus) — Adresse nicht geprüft"
        return f"smtp FAIL — {pr['detail'][:80]}"
    return f"smtp UNBEKANNT — {pr['detail'][:80]}"


def load_rows(wb) -> list[dict]:
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, i + 1).value for i, h in enumerate(headers)}
        if row.get("nr") is None:
            continue
        row["_row"] = r
        rows.append(row)
    return rows


def apply_research(rows: list[dict]) -> list[dict]:
    for row in rows:
        key = (int(row["nr"]), row["contact_type"])
        data = RESEARCH.get(key, {})
        row.update(data)
        if row["contact_type"] == "CEO/GF":
            row["email_body"] = build_ceo_body(row)
        else:
            row["email_body"] = build_eng_body(row)
        if row.get("status") == "routing_erforderlich":
            row["email_body_routing"] = routing_body(row)
    return rows


def run_smtp(rows: list[dict]) -> None:
    log = []
    for row in rows:
        primary = (row.get("email") or "").strip().lower()
        alts = parse_alt_emails(row.get("alt_email"))
        is_routing = primary.startswith("info@") or primary.startswith("office@") or row.get("status") == "routing_erforderlich"
        to_test = []
        if primary and primary not in to_test:
            to_test.append(primary)
        for a in alts:
            if a not in to_test:
                to_test.append(a)
        results = []
        for em in to_test:
            if em.startswith("info@") or em.startswith("office@"):
                continue
            results.append(smtp_check(em))
            time.sleep(0.2)
        label = smtp_label(primary, results, is_routing)
        row["smtp_check"] = label
        working = next((r for r in results if r["status"] == "ok"), None)
        if not is_routing and primary:
            pr = next((r for r in results if r["email"] == primary), None)
            if pr and pr["status"] == "fail" and working:
                row["email"] = working["email"]
                row["email_status"] = (row.get("email_status") or "") + " | SMTP-Alt"
                row["smtp_check"] = f"smtp FAIL primär — nutze {working['email']}"
        log.append({"nr": row["nr"], "name": row.get("name"), "email": row.get("email"), "smtp_check": label, "tested": results})
        print(f"  {row['company'][:30]:30} {row.get('name','')[:25]:25} {label}")
    return log


def write_excel(wb, rows: list[dict]) -> None:
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]
    if "email_body_routing" not in headers:
        headers.append("email_body_routing")
        ws.cell(1, len(headers), "email_body_routing").font = Font(bold=True)
    for row in rows:
        r = row["_row"]
        for h in headers:
            if h == "email_body_routing":
                val = row.get("email_body_routing", "")
            else:
                val = row.get(h, "")
            cell = ws.cell(r, headers.index(h) + 1, val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for h in ("email_body", "email_body_routing", "hook_fakt", "remarks"):
        if h in headers:
            ws.column_dimensions[get_column_letter(headers.index(h) + 1)].width = 48


def write_csv(rows: list[dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    extra = ["email_body_routing"] if any(r.get("email_body_routing") for r in rows) else []
    fields = HEADERS + [f for f in extra if f not in HEADERS]
    with CSV_OUT.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in fields})


def main() -> None:
    print("Lade Excel …")
    wb = openpyxl.load_workbook(XLSX)
    rows = load_rows(wb)
    print(f"Recherche anwenden ({len(rows)} Zeilen) …")
    rows = apply_research(rows)
    print("SMTP-Check …")
    log = run_smtp(rows)
    write_excel(wb, rows)
    try:
        wb.save(XLSX)
        print(f"Gespeichert: {XLSX}")
    except PermissionError:
        fallback = OUT_DIR / "260615_Kampagne_2_mit_Recherche.xlsx"
        wb.save(fallback)
        print(f"HINWEIS: Original gesperrt — gespeichert als {fallback}")
    write_csv(rows)
    print(f"CSV: {CSV_OUT}")
    (OUT_DIR / "_smtp_results_kampagne2.json").write_text(
        json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
