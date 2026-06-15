#!/usr/bin/env python3
"""Create Kontakte_Email sheet in 260615_Kampagne_2.xlsx for 14 target companies."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "260615_Kampagne_2.xlsx"
XLSX_FALLBACK = (
    ROOT / "04_Outreach/Kampagnen/2026-06-15_Kampagne_2/260615_Kampagne_2_mit_Kontakte.xlsx"
)
SHEET = "Kontakte_Email"

HEADERS = [
    "nr",
    "segment",
    "company",
    "location",
    "website",
    "contact_type",
    "name",
    "role",
    "email",
    "email_status",
    "email_pattern",
    "smtp_check",
    "source_url",
    "source_type",
    "alt_email",
    "central_email",
    "central_phone",
    "linkedin_url",
    "produkte_kurz",
    "zielkunden_kurz",
    "branche_kontext",
    "website_fakt",
    "hook_fakt",
    "hook_quelle",
    "hook_lens",
    "pain_kundensprache",
    "email_betreff",
    "email_body",
    "status",
    "remarks",
]

# fmt: off
COMPANIES = [
    # --- Intralogistik (Material_Handling) ---
    {
        "nr": 1, "segment": "Intralogistik",
        "company": "GEBHARDT Fördertechnik GmbH", "location": "DE, Sinsheim",
        "website": "https://www.gebhardt-foerdertechnik.de",
        "central_email": "info@gebhardt-foerdertechnik.de", "central_phone": "+49 7261 939-0",
        "email_domain": "gebhardt-foerdertechnik.de", "email_pattern": "{first}.{last}@gebhardt-foerdertechnik.de",
        "produkte_kurz": "Förderer, StoreBiter-Shuttle, AGV/AMR, Sortierer, WCS",
        "zielkunden_kurz": "Automotive, Food & Beverage, E-Commerce, 3PL",
        "branche_kontext": "Kundenspezifische Lager-/Fördersysteme; Layout-Varianten pro RFQ; typ. 4–10 Wochen Angebotszyklen (ETO).",
        "website_fakt": "3. Generation familiengeführt; Investition AMR/Shuttle — Website prüfen",
        "pain_kundensprache": "Parallele Layout-/Technologievarianten binden Senior-Ingenieure in der Kalkulation",
        "ceo_hint": "GF: Impressum/Management-Team recherchieren",
        "eng_hint": "Leiter Konstruktion / Engineering — LinkedIn + Impressum",
    },
    {
        "nr": 2, "segment": "Intralogistik",
        "company": "AUMUND Fördertechnik GmbH", "location": "DE, Rheinberg",
        "website": "https://www.aumund.com",
        "central_email": "info@aumund.com", "central_phone": "+49 2843 72-0",
        "email_domain": "aumund.com", "email_pattern": "{first}.{last}@aumund.com",
        "produkte_kurz": "Schwerlastförderer, Panzerketten, Elevatoren; Zement/Bergbau/Häfen",
        "zielkunden_kurz": "Zement, Bergbau, Hafen, Stahl, Chemie",
        "branche_kontext": "Großprojekt-ETO; lange Angebotszyklen; hohe Projektmargen-Risiken bei Varianten.",
        "website_fakt": "Nischenführer Bulk Conveying — Presse/News prüfen",
        "pain_kundensprache": "Großprojekt-Angebote: wo geht Engineering-Zeit — Layout oder Detailkalkulation?",
        "ceo_hint": "GF: Impressum recherchieren",
        "eng_hint": "Leiter Konstruktion / Projektleitung Engineering",
    },
    {
        "nr": 3, "segment": "Intralogistik",
        "company": "STIWA Group", "location": "AT, Attnang-Puchheim",
        "website": "https://www.stiwa.com",
        "central_email": "office@stiwa.com", "central_phone": "+43 7674 6600-0",
        "email_domain": "stiwa.com", "email_pattern": "{first}.{last}@stiwa.com",
        "produkte_kurz": "Montageautomation FAST-Zellen, Linear-Transfer, Material Handling",
        "zielkunden_kurz": "Automotive, Medical Devices, Electronics",
        "branche_kontext": "Sonderzellen pro Kunde; ETO-Montageautomation; AI/Automation-Investitionen.",
        "website_fakt": "Investition AI-basierte Montageautomation — News prüfen",
        "pain_kundensprache": "Jede Sonderzelle neu kalkuliert — Muster aus früheren Projekten nutzbar?",
        "ceo_hint": "GF: office@stiwa.com / Management",
        "eng_hint": "Leiter Entwicklung / Konstruktion",
    },
    {
        "nr": 4, "segment": "Intralogistik",
        "company": "DS Automotion GmbH", "location": "AT, Linz",
        "website": "https://www.ds-automotion.com",
        "central_email": "office@ds-automotion.com", "central_phone": "+43 732 7646-0",
        "email_domain": "ds-automotion.com", "email_pattern": "{first}.{last}@ds-automotion.com",
        "produkte_kurz": "Heavy-Duty AGV, Flottensteuerung; Automotive/Industrie",
        "zielkunden_kurz": "Automotive OEM, Industrie, Logistik",
        "branche_kontext": "AGV-Flotten ETO; SSI-Schäfer-Ökosystem; kundenspezifische Fahrzeugvarianten.",
        "website_fakt": "Schwerlast-AGV Linz — Website/Presse prüfen",
        "pain_kundensprache": "Kundenspezifische AGV-Flotten — Kostensicht der Varianten vor Freigabe?",
        "ceo_hint": "GF: Impressum recherchieren",
        "eng_hint": "CTO / Leiter Entwicklung AGV",
    },
    # --- Packaging Machinery ---
    {
        "nr": 5, "segment": "Packaging",
        "company": "Kolbus GmbH & Co. KG", "location": "DE, Rahden",
        "website": "https://www.kolbus.de",
        "central_email": "info@kolbus.de", "central_phone": "+49 5771 71-0",
        "email_domain": "kolbus.de", "email_pattern": "{first}.{last}@kolbus.de",
        "produkte_kurz": "Case-Maker Wellpappe, Buchbindung, Digital Workflow",
        "zielkunden_kurz": "Wellpappen-Converter, Print, Luxury Packaging",
        "branche_kontext": "Verpackungsmaschinenbau (V): PPWR, Formatwechsel, FMCG-Tender, RFQ nach Messe.",
        "website_fakt": "Post-Insolvenz 2019; Fokus Wellpappe/Print-Finishing — News prüfen",
        "pain_kundensprache": "Nach Restrukturierung: Kalkulationswissen institutionalisiert oder in Köpfen?",
        "ceo_hint": "GF post-Restrukturierung — Impressum/Handelsregister",
        "eng_hint": "Konstruktionsleiter Kolbus",
    },
    {
        "nr": 6, "segment": "Packaging",
        "company": "Windmöller & Hölscher KG", "location": "DE, Lengerich",
        "website": "https://www.wuh-group.com",
        "central_email": "info@wuh-group.com", "central_phone": "+49 5481 14-0",
        "email_domain": "wuh-group.com", "email_pattern": "{first}.{last}@wuh-group.com",
        "produkte_kurz": "Blown Film, Flexodruck, Beutelmaschinen, FILMATIC-Linien",
        "zielkunden_kurz": "Flexible Packaging Converter, FMCG, Filmhersteller",
        "branche_kontext": "Verpackung (V): Paper-based flexible packaging, PPWR, Materialumstellung Folie→Papier.",
        "website_fakt": "Technology Center Lengerich; Paper-based packaging — Website prüfen",
        "pain_kundensprache": "Materialumstellung im Angebot — Kostenauswirkung der Maschinenvariante früh sichtbar?",
        "ceo_name": "Hendrik Nienkemper", "ceo_role": "Sprecher der Geschäftsführung",
        "ceo_linkedin": "https://www.linkedin.com/in/hendrik-nienkemper",
        "eng_hint": "Leiter Entwicklung W&H Lengerich",
    },
    {
        "nr": 7, "segment": "Packaging",
        "company": "Theegarten-Pactec GmbH & Co. KG", "location": "DE, Dresden",
        "website": "https://www.theegarten-pactec.com",
        "central_email": "info@theegarten-pactec.com", "central_phone": "+49 351 28803-0",
        "email_domain": "theegarten-pactec.com", "email_pattern": "{first}.{last}@theegarten-pactec.com",
        "produkte_kurz": "High-Speed Süßwaren-Wrapper, Schokolade, Servo-Plattformen",
        "zielkunden_kurz": "Ferrero, Lindt, Haribo, Mondelez, Mars",
        "branche_kontext": "Verpackung (V): Export, Formatvarianten, RFQ-Wellen nach Kapazitätsausbau.",
        "website_fakt": "Produktionshalle Dresden 2021; Exportmärkte — News prüfen",
        "pain_kundensprache": "Parallele Süßwaren-RFQs nach Kapazitätsausbau — Engineering in Kalkulation?",
        "ceo_name": "Thomas Rother", "ceo_role": "Geschäftsführer",
        "eng_hint": "Leiter Konstruktion Theegarten-Pactec",
    },
    {
        "nr": 8, "segment": "Packaging",
        "company": "Optima Packaging Group GmbH", "location": "DE, Schwäbisch Hall",
        "website": "https://www.optima-packaging.com",
        "central_email": "info@optima-packaging.com", "central_phone": "+49 791 9495-0",
        "email_domain": "optima-packaging.com", "email_pattern": "{first}.{last}@optima-packaging.com",
        "produkte_kurz": "Pharma/Cosmetics/Consumer-Linien, INOVA, Serialization, Digital Twin",
        "zielkunden_kurz": "Pfizer, Roche, AstraZeneca, L'Oréal, P&G, Henkel",
        "branche_kontext": "Verpackung (V): GMP + PPWR; Engineering Change Orders; Tribal Knowledge.",
        "website_fakt": "INOVA-Akquisition; Digital Twin/IIoT — Presse prüfen",
        "pain_kundensprache": "GMP-Varianten und Compliance-Änderungen vor Freigabe in belastbare Kosten?",
        "ceo_name": "Hans Bühler", "ceo_role": "Geschäftsführer",
        "eng_hint": "Leiter Konstruktion Optima Packaging",
    },
    {
        "nr": 9, "segment": "Packaging",
        "company": "Gerhard Schubert GmbH", "location": "DE, Crailsheim",
        "website": "https://www.schubert.group",
        "central_email": "info@schubert.group", "central_phone": "+49 7951 400-0",
        "email_domain": "gerhard-schubert.de", "email_pattern": "{f}.{last}@gerhard-schubert.de",
        "produkte_kurz": "TLM Pick-and-Place, F4-Roboter, Schubert-Pharma, Vision/AI",
        "zielkunden_kurz": "Food, Pharma, Cosmetics, Medical Devices",
        "branche_kontext": "Verpackung (V): Modulare Linien, hohe Variantenkomplexität, Digital Twin.",
        "website_fakt": "Schubert-Pharma Expansion; AI Vision — Website prüfen",
        "pain_kundensprache": "Modulare Linien = Variantenvielfalt — RFQ-Tempo vs. Engineering-Kapazität?",
        "ceo_name": "Ralf Schubert", "ceo_role": "Geschäftsführer (Technologie-Division)",
        "ceo_email": "r.schubert@gerhard-schubert.de", "ceo_status": "abgeleitet (hoch)",
        "ceo_source": "https://www.schubert.group/en/imprint/", "ceo_source_type": "impressum",
        "eng_name": "Marcus Schindler", "eng_role": "Bereichsleiter Konstruktion",
        "eng_email": "m.schindler@gerhard-schubert.de", "eng_status": "abgeleitet (hoch)",
        "eng_source": "https://www.schubert.group/en/group/locations/", "eng_source_type": "firmenwebsite",
        "eng_linkedin": "https://www.linkedin.com/in/marcus-schindler-789554131",
        "alt_email": "t.neumann@gerhard-schubert.de (Pattern-Bestätigung)",
    },
    {
        "nr": 10, "segment": "Packaging",
        "company": "Bielomatik Leuze GmbH + Co. KG", "location": "DE, Neuffen",
        "website": "https://www.bielomatik.com",
        "central_email": "info@bielomatik.com", "central_phone": "+49 7025 12-0",
        "email_domain": "bielomatik.com", "email_pattern": "{first}.{last}@bielomatik.com",
        "produkte_kurz": "Paper Converting, Friction Welding, Tissue/Office Paper",
        "zielkunden_kurz": "Mondi, Sappi, Tissue, Automotive (Friction Welding)",
        "branche_kontext": "Verpackung/Paper (V): Paper-Converting unter Verpackungsdruck; ETO-Varianten.",
        "website_fakt": "Diversifikation Paper + Friction Welding — News prüfen",
        "pain_kundensprache": "Varianten im Paper-Converting — frühe Kostensicht im Design?",
        "ceo_hint": "GF: Impressum recherchieren",
        "eng_hint": "Leiter Konstruktion Neuffen",
    },
    {
        "nr": 11, "segment": "Packaging",
        "company": "EREMA Engineering Recycling Maschinen", "location": "AT, Ansfelden",
        "website": "https://www.erema.com",
        "central_email": "office@erema.at", "central_phone": "+43 732 3190-0",
        "email_domain": "erema.com", "email_pattern": "{first}.{last}@erema.com",
        "produkte_kurz": "INTAREMA/VACUREMA PCR-Recycling, Film-Linien, PURE LOOP",
        "zielkunden_kurz": "Packaging Producer, Recycler, FMCG mit PCR-Zielen",
        "branche_kontext": "Verpackung/Recycling (V): Circular Economy, PPWR, PCR-Anteil als Kundenspec.",
        "website_fakt": "Marktführer Plastic Recycling; PPWR-Treiber — Presse prüfen",
        "pain_kundensprache": "PCR-Anteil als Kundenspec — Maschinenvariante im Design kalkulierbar?",
        "ceo_name": "Manfred Hackl", "ceo_role": "CEO",
        "eng_hint": "Leiter Entwicklung / CTO EREMA",
    },
    {
        "nr": 12, "segment": "Packaging",
        "company": "Bobst Group SA", "location": "CH, Mex",
        "website": "https://www.bobst.com",
        "central_email": "info@bobst.com", "central_phone": "",
        "email_domain": "bobst.com", "email_pattern": "{first}.{last}@bobst.com",
        "produkte_kurz": "Folding Carton, Corrugated, Flexible Packaging Machinery",
        "zielkunden_kurz": "Converter, Brand Owner, Global Packaging",
        "branche_kontext": "Verpackung (V) Konzern: Mix ETO/Standard; Konfigurationsvielfalt; Presse-Routing wahrscheinlich.",
        "website_fakt": "Konzern — Sustainability/Digital auf bobst.com prüfen",
        "pain_kundensprache": "Konfigurationsvielfalt — Bottleneck Angebot vs. Engineering?",
        "ceo_hint": "CEO über Presse/Management-Seite; ggf. press@bobst.com Routing",
        "eng_hint": "Head of Engineering / CTO — Konzern-Struktur beachten",
        "remarks_company": "Lead B · Konzern · Presse-Routing prüfen (05_Konzern_ueber_Presse.md)",
    },
    {
        "nr": 13, "segment": "Packaging",
        "company": "Christian Senning Verpackungsmaschinen GmbH", "location": "DE, Bremen",
        "website": "https://www.senning.de",
        "central_email": "info@senning.de", "central_phone": "+49 421 55 90 0",
        "email_domain": "senning.de", "email_pattern": "{first}.{last}@senning.de",
        "produkte_kurz": "Süßwaren/Tobacco/Food Wrapper, ETO",
        "zielkunden_kurz": "Confectionery, Tobacco, Food",
        "branche_kontext": "Verpackung (V): Inhaber-geführt ~50 MA; schnelle Entscheidungen; jede Anfrage Unikat.",
        "website_fakt": "Familienunternehmen seit 1949 — Impressum prüfen",
        "pain_kundensprache": "Kleiner ETO — Kalkulation am GF/Ingenieur gebunden?",
        "ceo_hint": "GF/Inhaber — Impressum senning.de",
        "eng_hint": "Konstruktionsleitung",
    },
    {
        "nr": 14, "segment": "Packaging",
        "company": "Fill GmbH (packaging division)", "location": "AT, Gurten",
        "website": "https://www.fill.co.at",
        "central_email": "info@fill.co.at", "central_phone": "",
        "email_domain": "fill.co.at", "email_pattern": "k.A. — Pattern recherchieren",
        "produkte_kurz": "Sondermaschinenbau inkl. Verpackungslinien (Sparte klären)",
        "zielkunden_kurz": "Industrie, ggf. Packaging — Website prüfen",
        "branche_kontext": "Verpackung/ETO Mix (V+E): Sondermaschinenbau breit; Verpackungssparte identifizieren.",
        "website_fakt": "fill.co.at — Packaging-Division auf Website lokalisieren",
        "pain_kundensprache": "ETO-Projektgeschäft — wo stocken RFQ-Schleifen?",
        "ceo_hint": "GF Fill — Impressum",
        "eng_hint": "Leiter Konstruktion Verpackungssparte",
        "remarks_company": "Lead B · Verpackungssparte auf Website verifizieren",
    },
]
# fmt: on


def hook_lens_ceo(segment: str) -> str:
    return "GF_kapazitaet_marge"


def hook_lens_eng(segment: str) -> str:
    return "Eng_varianten_freigabe"


def draft_hook_fakt(c: dict, role: str) -> str:
  """Draft hook — website_fakt must be verified/updated during research."""
  seg = c["segment"]
  fact = c.get("website_fakt", "")
  pain = c.get("pain_kundensprache", "")
  if role == "CEO/GF":
      if seg == "Packaging":
          return (
              f"{fact} Im Verpackungsmaschinenbau höre ich oft, dass FMCG-Specs und "
              f"Nachhaltigkeitsthemen früher in die Angebotsphase rutschen — {pain}"
          )
      return (
          f"{fact} Bei kundenspezifischen Förder- und Lagersystemen bindet die "
          f"Variantenkalkulation oft Senior-Ingenieure — {pain}"
      )
  if seg == "Packaging":
      return (
          f"{fact} Bei Sondermaschinen für Verpackungslinien frage ich mich, ob "
          f"Material- und Formatvarianten bei Ihnen eher vor oder nach der Freigabe "
          f"in derselben Kostensicht bewertet werden."
      )
  return (
      f"{fact} Bei ETO-Intralogistik-Projekten höre ich häufig, dass Layout- und "
      f"Technologievarianten parallel laufen — wo stockt bei Ihnen die Bewertung in Euro?"
  )


def draft_betreff(c: dict, role: str) -> str:
    short = c["company"].split()[0]
    if role == "CEO/GF":
        return f"Kurze Frage zu {short} — Angebotskapazität oder Marge?"
    return f"{short} — Varianten vor Freigabe?"


def build_rows() -> list[dict]:
    rows = []
    for c in COMPANIES:
        base = {
            "nr": c["nr"],
            "segment": c["segment"],
            "company": c["company"],
            "location": c["location"],
            "website": c["website"],
            "central_email": c["central_email"],
            "central_phone": c.get("central_phone", ""),
            "produkte_kurz": c["produkte_kurz"],
            "zielkunden_kurz": c["zielkunden_kurz"],
            "branche_kontext": c["branche_kontext"],
            "website_fakt": c["website_fakt"],
            "pain_kundensprache": c["pain_kundensprache"],
            "email_pattern": c["email_pattern"],
            "status": "recherche_offen",
            "remarks": c.get("remarks_company", ""),
        }
        # CEO row
        ceo = {**base, "contact_type": "CEO/GF"}
        ceo["name"] = c.get("ceo_name", "")
        ceo["role"] = c.get("ceo_role", c.get("ceo_hint", "recherchieren"))
        ceo["email"] = c.get("ceo_email", "")
        ceo["email_status"] = c.get("ceo_status", "nicht recherchiert")
        ceo["source_url"] = c.get("ceo_source", "")
        ceo["source_type"] = c.get("ceo_source_type", "")
        ceo["linkedin_url"] = c.get("ceo_linkedin", "")
        ceo["alt_email"] = c.get("alt_email", c["central_email"])
        ceo["hook_lens"] = hook_lens_ceo(c["segment"])
        ceo["hook_fakt"] = draft_hook_fakt(c, "CEO/GF")
        ceo["hook_quelle"] = c.get("ceo_source", c["website"] + " — bitte aktualisieren")
        ceo["email_betreff"] = draft_betreff(c, "CEO/GF")
        ceo["email_body"] = ""
        ceo["smtp_check"] = ""
        if not ceo["name"]:
            ceo["remarks"] = (ceo["remarks"] + " | " if ceo["remarks"] else "") + c.get("ceo_hint", "")
        rows.append(ceo)

        # Engineering row
        eng = {**base, "contact_type": "Engineering"}
        eng["name"] = c.get("eng_name", "")
        eng["role"] = c.get("eng_role", c.get("eng_hint", "recherchieren"))
        eng["email"] = c.get("eng_email", "")
        eng["email_status"] = c.get("eng_status", "nicht recherchiert")
        eng["source_url"] = c.get("eng_source", "")
        eng["source_type"] = c.get("eng_source_type", "")
        eng["linkedin_url"] = c.get("eng_linkedin", "")
        eng["alt_email"] = c.get("alt_email", c["central_email"])
        eng["hook_lens"] = hook_lens_eng(c["segment"])
        eng["hook_fakt"] = draft_hook_fakt(c, "Engineering")
        eng["hook_quelle"] = c.get("eng_source", c["website"] + " — bitte aktualisieren")
        eng["email_betreff"] = draft_betreff(c, "Engineering")
        eng["email_body"] = ""
        eng["smtp_check"] = ""
        if not eng["name"]:
            eng["remarks"] = (eng["remarks"] + " | " if eng["remarks"] else "") + c.get("eng_hint", "")
        rows.append(eng)
    return rows


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    seg_pack = PatternFill("solid", fgColor="E2EFDA")
    seg_intra = PatternFill("solid", fgColor="DDEBF7")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
  # column widths
    widths = {
        "company": 36, "hook_fakt": 50, "email_body": 40, "branche_kontext": 40,
        "website_fakt": 35, "produkte_kurz": 32, "remarks": 30,
    }
    for col, h in enumerate(HEADERS, 1):
        w = widths.get(h, 14)
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(2, ws.max_row + 1):
        seg = ws.cell(r, HEADERS.index("segment") + 1).value
        fill = seg_pack if seg == "Packaging" else seg_intra
        for c in range(1, len(HEADERS) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(r, c).fill = fill


def write_sheet(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    for r, row in enumerate(rows, 2):
        for c, key in enumerate(HEADERS, 1):
            ws.cell(row=r, column=c, value=row.get(key, ""))
    style_sheet(ws)


def main() -> None:
    rows = build_rows()
    try:
        wb = openpyxl.load_workbook(XLSX)
        write_sheet(wb, rows)
        wb.save(XLSX)
        print(f"OK: {SHEET} — {len(rows)} Zeilen in {XLSX}")
    except PermissionError:
        import shutil

        shutil.copy2(XLSX, XLSX_FALLBACK)
        wb = openpyxl.load_workbook(XLSX_FALLBACK)
        write_sheet(wb, rows)
        wb.save(XLSX_FALLBACK)
        print(
            f"HINWEIS: {XLSX.name} ist gesperrt (Excel schließen).\n"
            f"OK: {SHEET} — {len(rows)} Zeilen in {XLSX_FALLBACK}"
        )


if __name__ == "__main__":
    main()
