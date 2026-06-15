#!/usr/bin/env python3
"""Finalize Kampagne 2 outreach: deep research, email variants, SMTP, Excel table export."""
from __future__ import annotations

import csv
import json
import re
import smtplib
import socket
import time
import unicodedata
from copy import deepcopy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "04_Outreach/Kampagnen/2026-06-15_Kampagne_2"
MASTER = ROOT / "260615_Kampagne_2.xlsx"
KONTAKTE_COPY = OUT_DIR / "260615_Kampagne_2_mit_Kontakte.xlsx"
CSV_OUT = OUT_DIR / "260615_Kontakte_Email.csv"
XLSX_OUT = OUT_DIR / "260615_Kontakte_Email.xlsx"
SMTP_LOG = OUT_DIR / "_smtp_results_kampagne2.json"
SHEET = "Kontakte_Email"
FROM_ADDR = "verify@loopforgelab.com"
TIMEOUT = 8
MAX_VARIANTS_SMTP = 5
SMTP_CACHE: dict[str, dict] = {}

if SMTP_LOG.exists():
    try:
        for entry in json.loads(SMTP_LOG.read_text(encoding="utf-8")):
            for t in entry.get("tested", []):
                if t.get("email"):
                    SMTP_CACHE[t["email"].lower()] = t
    except Exception:
        pass

HEADERS = [
    "nr", "segment", "company", "location", "website", "contact_type", "name", "role",
    "email", "email_status", "email_pattern", "smtp_check", "email_variants_tested",
    "source_url", "source_type", "alt_email", "central_email", "central_phone",
    "linkedin_url", "produkte_kurz", "zielkunden_kurz", "branche_kontext",
    "website_fakt", "hook_fakt", "hook_quelle", "hook_lens", "pain_kundensprache",
    "email_betreff", "email_body", "email_body_routing", "status", "remarks",
]


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().replace("ß", "ss").strip()


def parse_name(full: str) -> tuple[str, str]:
    clean = re.sub(r"^(Dr\.-Ing\.|Dr\.|DI|Prof\.)\s*", "", full.strip())
    parts = clean.split()
    if len(parts) >= 2:
        return parts[0], parts[-1]
    return clean, clean


def build_candidates(first: str, last: str, domain: str, extras: list[str] | None = None) -> list[str]:
    fi, la = norm(first), norm(last)
    Fi, La = first.strip().split()[0] if first else "", last.strip()
    Fi = Fi[0].upper() + Fi[1:] if Fi else ""
    La = La[0].upper() + La[1:] if La else ""
    out = [
        f"{Fi}.{La}@{domain}",
        f"{fi}.{la}@{domain}",
        f"{fi[0]}.{la}@{domain}" if fi else "",
        f"{Fi[0]}.{La}@{domain}" if Fi else "",
        f"{fi}{la}@{domain}",
        f"{Fi}{La}@{domain}",
        f"{fi}_{la}@{domain}",
        f"{fi}-{la}@{domain}",
    ]
    if extras:
        out.extend(extras)
    seen, uniq = set(), []
    for e in out:
        e = e.strip().lower()
        if e and "@" in e and e not in seen:
            seen.add(e)
            uniq.append(e)
    return uniq


# fmt: off
CONTACTS: list[dict] = [
    # === 1 GEBHARDT ===
    {"nr": 1, "segment": "Intralogistik", "company": "GEBHARDT Fördertechnik GmbH", "location": "DE, Sinsheim",
     "website": "https://www.gebhardt-group.com", "central_email": "info@gebhardt-group.com", "central_phone": "+49 7261 939-0",
     "email_domain": "gebhardt-group.com", "pattern_hint": "{f}.{last}@gebhardt-group.com (s.glaesner verifiziert)",
     "produkte_kurz": "Förderer, StoreBiter-Shuttle, AGV/AMR, Sortierer, WCS",
     "zielkunden_kurz": "Automotive, Food & Beverage, E-Commerce, 3PL",
     "branche_kontext": "Kundenspezifische Lager-/Fördersysteme; Layout-Varianten pro RFQ; 4–10 Wochen Angebotszyklen (ETO).",
     "pain_kundensprache": "Parallele Layout-/Technologievarianten binden Senior-Ingenieure in der Kalkulation",
     "contacts": [
         {"type": "CEO/GF", "name": "Marco Gebhardt", "role": "Geschäftsführer / Inhaber (3. Generation)",
          "domain": "gebhardt-group.com", "extras": [],
          "source_url": "https://pl.gebhardt-group.com/index.php/pl/imprint.html", "source_type": "impressum",
          "website_fakt": "LogiMAT 2026; US-Produktionsausbau Streetsboro (Okt 2025); familiengeführt seit 1952.",
          "hook_fakt": "Mit dem US-Manufacturing-Ausbau und LogiMAT 2026 laufen bei Ihnen vermutlich mehr Layout- und Shuttle-Varianten parallel in der Angebotsphase — typisch für ETO-Intralogistik.",
          "hook_quelle": "https://us.gebhardt-group.com/company/news/gebhardt-intralogistics-north-america-names-ralf-buerkle-as-chief-operating-officer.html",
          "betreff": "LogiMAT + US-Ausbau — bindet Engineering Ihre Angebotskapazität?"},
         {"type": "Engineering", "name": "Andreas Hooge", "role": "Head of Project Engineering",
          "domain": "gebhardt-group.com", "extras": [],
          "source_url": "https://www.linkedin.com/company/gebhardt-group", "source_type": "linkedin",
          "website_fakt": "Neuer Retrofit-Geschäftsbereich; Project Engineering + Project Design in Sinsheim.",
          "hook_fakt": "Retrofit und Warehouse-Robotics erweitern das Portfolio — ich frage mich, ob Layout- und Steuerungsvarianten in derselben RFQ-Phase in Euro bewertet werden.",
          "hook_quelle": "https://jobs.gebhardt-group.com/",
          "betreff": "Retrofit & Robotics — wo stocken Varianten vor Freigabe?"},
     ]},
    # === 2 AUMUND ===
    {"nr": 2, "segment": "Intralogistik", "company": "AUMUND Fördertechnik GmbH", "location": "DE, Rheinberg",
     "website": "https://www.aumund.com", "central_email": "info@aumund.com", "central_phone": "+49 2843 72-0",
     "email_domain": "aumund.com", "pattern_hint": "{first}.{last}@aumund.com",
     "produkte_kurz": "Schwerlastförderer, Panzerketten, Elevatoren; Zement/Bergbau/Häfen",
     "zielkunden_kurz": "Zement, Bergbau, Hafen, Stahl, Chemie",
     "branche_kontext": "Großprojekt-ETO; lange Angebotszyklen; hohe Projektmargen-Risiken bei Varianten.",
     "pain_kundensprache": "Großprojekt-Angebote: wo geht Engineering-Zeit — Layout oder Detailkalkulation?",
     "contacts": [
         {"type": "CEO/GF", "name": "Dr.-Ing. Pietro de Michieli", "role": "CEO / Geschäftsführer",
          "domain": "aumund.com", "primary_email": "pietro.demichieli@aumund.com", "extras": [],
          "source_url": "https://www.aumund.com/en/company/management/", "source_type": "firmenwebsite",
          "website_fakt": "Nischenführer Bulk-Material-Handling; Großprojekte weltweit in Zement/Bergbau.",
          "hook_fakt": "Bei Großprojekt-ETO mit langen Angebotszyklen frage ich mich, ob Layout- oder Detailvarianten Ihre Senior-Ingenieure stärker binden.",
          "hook_quelle": "https://www.aumund.com",
          "betreff": "Großprojekt-Kalkulation bei AUMUND — Layout oder Detail?"},
         {"type": "Engineering", "name": "Reiner Furthmann", "role": "CTO",
          "domain": "aumund.com", "extras": [],
          "source_url": "https://www.aumund.com/en/company/management/", "source_type": "firmenwebsite",
          "website_fakt": "Schwerlast- und Bulk-Anlagen; variantenreiche Konfigurationen pro Projekt.",
          "hook_fakt": "Bei Schwerlast-ETO höre ich oft: Die teuersten Überraschungen entstehen vor der Freigabe — wenn Layout- und Detailvarianten nicht in derselben Kostensicht laufen.",
          "hook_quelle": "https://www.aumund.com/en/company/management/",
          "betreff": "Schwerlast-Varianten — Nacharbeit oder Durchlaufzeit?"},
     ]},
    # === 3 STIWA ===
    {"nr": 3, "segment": "Intralogistik", "company": "STIWA Group", "location": "AT, Attnang-Puchheim",
     "website": "https://www.stiwa.com", "central_email": "office@stiwa.com", "central_phone": "+43 7674 6600-0",
     "email_domain": "stiwa.com", "pattern_hint": "{first}.{last}@stiwa.com (Website verifiziert)",
     "produkte_kurz": "FAST-Montagezellen, Linear-Transfer, Material Handling",
     "zielkunden_kurz": "Automotive, Medical Devices, Electronics",
     "branche_kontext": "Sonderzellen pro Kunde; ETO-Montageautomation.",
     "pain_kundensprache": "Jede Sonderzelle neu kalkuliert — Muster aus früheren Projekten nutzbar?",
     "contacts": [
         {"type": "CEO/GF", "name": "Michael Fuchshuber", "role": "CEO STIWA Automation",
          "domain": "stiwa.com", "extras": ["michael.fuchshuber@stiwa.com"],
          "source_url": "https://www.stiwa.com/en/manufacturing/contact", "source_type": "firmenwebsite",
          "website_fakt": "FAST-Montagezellen; KI-gestützte Montageautomation; ~265 MA in AT.",
          "hook_fakt": "Jede Sonderzelle ist ein Unikat — rutschen frühere Projektdaten bei neuen RFQs wirklich in die Kalkulation?",
          "hook_quelle": "https://www.stiwa.com",
          "betreff": "Sonderzellen-Kalkulation bei STIWA — Daten oder Neuaufbau?"},
         {"type": "Engineering", "name": "Gerhard Berer", "role": "Bereichsleiter Engineering",
          "domain": "stiwa.com", "extras": ["gerhard.berer@stiwa.com"],
          "source_url": "https://www.linkedin.com/in/gerhard-berer-656370150", "source_type": "linkedin",
          "website_fakt": "Engineering für Montageautomation; Bereichsleitung seit 2010.",
          "hook_fakt": "Bei FAST-Zellen: Wo verlieren Sie mehr Zeit — Varianten durchrechnen oder auf Einkauf/Kalkulation warten?",
          "hook_quelle": "https://www.linkedin.com/in/gerhard-berer-656370150",
          "betreff": "FAST-Zellen — Varianten vor Freigabe bei STIWA?"},
     ]},
    # === 4 DS Automotion ===
    {"nr": 4, "segment": "Intralogistik", "company": "DS Automotion GmbH", "location": "AT, Linz",
     "website": "https://www.ds-automotion.com", "central_email": "info@ds-automotion.com", "central_phone": "+43 732 7646-0",
     "email_domain": "ds-automotion.com", "pattern_hint": "{first}.{last}@ds-automotion.com",
     "produkte_kurz": "Heavy-Duty AGV, Flottensteuerung; Automotive/Industrie",
     "zielkunden_kurz": "Automotive OEM, Industrie, Logistik",
     "branche_kontext": "AGV-Flotten ETO; SSI-Schäfer-Ökosystem; kundenspezifische Fahrzeugvarianten.",
     "pain_kundensprache": "Kundenspezifische AGV-Flotten — Kostensicht der Varianten vor Freigabe?",
     "contacts": [
         {"type": "CEO/GF", "name": "Wolfgang Hillinger", "role": "Geschäftsführer",
          "domain": "ds-automotion.com", "extras": [],
          "source_url": "https://www.ds-automotion.com/en-us/imprint/", "source_type": "impressum",
          "website_fakt": "SSI-Schäfer-Tochter; Schwerlast-AGV aus Linz; Navigation/Steuerung in-house.",
          "hook_fakt": "Kundenspezifische AGV-Flotten mit variantenreicher Konfiguration — limitiert das Angebotsvolumen Kapazität oder Marge?",
          "hook_quelle": "https://www.ds-automotion.com",
          "betreff": "AGV-Flotten-RFQs — Kapazität oder Marge bei DS Automotion?"},
         {"type": "Engineering", "name": "Thomas Gschwentenwein", "role": "CTO / Bereichsleiter Technologie & Entwicklung",
          "domain": "ds-automotion.com", "extras": ["thomas.gschwentenwein@ds-automotion.com"],
          "source_url": "https://www.ds-automotion.com/en-us/company/about-us/", "source_type": "firmenwebsite",
          "website_fakt": "Schwerlast-AGV-Entwicklung; kundenspezifische Fahrzeug- und Flottenvarianten.",
          "hook_fakt": "Bei AGV-ETO: Trifft Nacharbeit nach Freigabe zu — oder stockt die Durchlaufzeit der Varianten in der Konstruktion?",
          "hook_quelle": "https://www.ds-automotion.com",
          "betreff": "AGV-Varianten — Freigabe oder Nacharbeit bei Ihnen?"},
     ]},
    # === 5 Kolbus ===
    {"nr": 5, "segment": "Packaging", "company": "KOLBUS Group GmbH", "location": "DE, Rahden",
     "website": "https://www.kolbus.de", "central_email": "info@kolbus.de", "central_phone": "+49 5771 71-0",
     "email_domain": "kolbus.de", "pattern_hint": "{First}.{Last}@kolbus.de (Markus.Dammann verifiziert)",
     "produkte_kurz": "Case-Maker Wellpappe, Buchbindung, Digital Workflow",
     "zielkunden_kurz": "Wellpappen-Converter, Print, Luxury Packaging",
     "branche_kontext": "Verpackungsmaschinenbau (V): PPWR, Formatwechsel, FMCG-Tender.",
     "pain_kundensprache": "Nach Restrukturierung: Kalkulationswissen institutionalisiert oder in Köpfen?",
     "contacts": [
         {"type": "CEO/GF", "name": "Michael R. Bach", "role": "CEO (seit Jan 2026)",
          "domain": "kolbus.de", "extras": ["Michael.Bach@kolbus.de", "michael.bach@kolbus.de", "M.R.Bach@kolbus.de"],
          "source_url": "https://kolbus.com/de/unternehmen/news/news/new-management-team-strengthens-kolbus-group-gmbh-in-rahden",
          "source_type": "presse",
          "website_fakt": "Neues 4-köpfiges GL-Team seit Jan 2026; Max Valier Holding; Fokus Wellpappe/Print.",
          "hook_fakt": "Mit neuem CEO und frischem GL-Team: Ist Kalkulationswissen für Wellpappen-Varianten institutionalisiert — oder noch in Köpfen gebunden?",
          "hook_quelle": "https://kolbus.com/de/unternehmen/news/",
          "betreff": "Neues GL-Team Kolbus — wo sitzt Kalkulationswissen?"},
         {"type": "Engineering", "name": "Mathias Sieckmann", "role": "CCO & CTO (Geschäftsleitung)",
          "domain": "kolbus.de", "extras": ["Mathias.Sieckmann@kolbus.de", "mathias.sieckmann@kolbus.de"],
          "source_url": "https://www.thepackagingportal.com/industry-news/new-senior-management-team-at-kolbus/",
          "source_type": "presse",
          "website_fakt": "Case-Maker und Buchbinderei; variantenreiche Maschinenkonfigurationen.",
          "hook_fakt": "Bei Wellpappen-/Print-Maschinen: Werden Format- und Materialvarianten vor oder nach der Freigabe in derselben Kostensicht bewertet?",
          "hook_quelle": "https://kolbus.com",
          "betreff": "Wellpappen-Varianten — vor oder nach Freigabe kalkuliert?"},
     ]},
    # === 6 W&H ===
    {"nr": 6, "segment": "Packaging", "company": "Windmöller & Hölscher KG", "location": "DE, Lengerich",
     "website": "https://www.wh.group", "central_email": "info@wuh-group.com", "central_phone": "+49 5481 14-0",
     "email_domain": "wuh-group.com", "pattern_hint": "{First}.{Last}@wuh-group.com",
     "produkte_kurz": "Blown Film, Flexodruck, Beutelmaschinen, FILMATIC-Linien",
     "zielkunden_kurz": "Flexible Packaging Converter, FMCG, Filmhersteller",
     "branche_kontext": "Verpackung (V): Paper-based flexible packaging, PPWR, Folie→Papier.",
     "pain_kundensprache": "Materialumstellung im Angebot — Kostenauswirkung der Maschinenvariante früh sichtbar?",
     "contacts": [
         {"type": "CEO/GF", "name": "Dr. Falco Paepenmüller", "role": "CEO (vormals CTO)",
          "domain": "wuh-group.com", "primary_email": "falco.paepenmueller@wuh-group.com", "extras": [],
          "source_url": "https://www.wh.group/de/unternehmen/w_h_group/", "source_type": "firmenwebsite",
          "website_fakt": "Paper-based flexible packaging (PPWR); Technology Center Lengerich; ~3.400 MA.",
          "hook_fakt": "Materialumstellung Folie→Papier treibt neue Maschinenvarianten — rutschen FMCG-Specs und PPWR früher in Ihre Angebotsphase?",
          "hook_quelle": "https://www.wh.group/de/unternehmen/w_h_group/",
          "betreff": "Folie→Papier & PPWR — früher in der W&H-Kalkulation?"},
         {"type": "Engineering", "name": "Dr. Björn Feldhaus", "role": "Director R&D & Sustaining Engineering",
          "domain": "wuh-group.com", "primary_email": "bjoern.feldhaus@wuh-group.com", "extras": [],
          "source_url": "https://www.linkedin.com/in/dr-bj%C3%B6rn-feldhaus-4b4327153", "source_type": "linkedin",
          "website_fakt": "R&D Flexodruck/Extrusion; technotrans-Partnerschaft Farbversorgung Flexodruck.",
          "hook_fakt": "Bei Flexo-/Extrusions-Varianten: Startet jede Sonderanfrage trotz vorhandener CAD-Daten fast bei Null?",
          "hook_quelle": "https://www.wirtschaft-aktuell.de/news/technotrans-und-wh-vertiefen-partnerschaft",
          "betreff": "Flexo-R&D — Datenschatz oder Neuaufbau pro RFQ?"},
     ]},
    # === 7 Theegarten ===
    {"nr": 7, "segment": "Packaging", "company": "Theegarten-Pactec GmbH & Co. KG", "location": "DE, Dresden",
     "website": "https://www.theegarten-pactec.de", "central_email": "pactec@theegarten-pactec.de", "central_phone": "+49 351 2573 0",
     "email_domain": "theegarten-pactec.de", "pattern_hint": "{f}.{last}@theegarten-pactec.de (d.schibur verifiziert)",
     "produkte_kurz": "High-Speed Süßwaren-Wrapper, Servo-Plattformen, FPC6",
     "zielkunden_kurz": "Ferrero, Lindt, Mars, Mondelez, FMCG Confectionery",
     "branche_kontext": "Verpackung (V): Export 95%, Formatvarianten, RFQ nach interpack.",
     "pain_kundensprache": "Parallele Süßwaren-RFQs — Engineering in der Kalkulation?",
     "contacts": [
         {"type": "CEO/GF", "name": "Markus Rustler", "role": "Geschäftsführer / interpack President",
          "domain": "theegarten-pactec.de", "extras": ["markus.rustler@theegarten-pactec.de", "m.rustler@theegarten-pactec.de"],
          "source_url": "https://www.theegarten-pactec.de/en/impressum/", "source_type": "impressum",
          "website_fakt": "interpack 2026 FPC6 Flow; Robert Berge verstärkt GL ab Jun 2026; Nachfolge Röhm 2028.",
          "hook_fakt": "Nach interpack 2026 und der FPC6-Vorstellung laufen bei Ihnen vermutlich parallele Süßwaren-RFQs — bindet Variantenkalkulation Senior-Ingenieure?",
          "hook_quelle": "https://packaging-journal.de/en/theegarten-pactec-expands-its-management-team/",
          "betreff": "Nach interpack — RFQ-Welle und Engineering-Kapazität?"},
         {"type": "Engineering", "name": "Dr. Egbert Röhm", "role": "Geschäftsführer Technologie & Produktion",
          "domain": "theegarten-pactec.de", "primary_email": "e.roehm@theegarten-pactec.de",
          "extras": ["egbert.roehm@theegarten-pactec.de"],
          "source_url": "https://www.theegarten-pactec.de/en/impressum/", "source_type": "impressum",
          "website_fakt": "GF Technologie/Produktion; strukturierte Nachfolge bis 2028; Export ~95%.",
          "hook_fakt": "Bei High-Speed-Linien: Wo verlieren Sie mehr Zeit — Formatvarianten durchrechnen oder auf Kalkulation warten?",
          "hook_quelle": "https://www.confectioneryproduction.com/news/58097/",
          "betreff": "High-Speed-Varianten — Durchlaufzeit oder Nacharbeit?"},
     ]},
    # === 8 Optima ===
    {"nr": 8, "segment": "Packaging", "company": "Optima Packaging Group GmbH", "location": "DE, Schwäbisch Hall",
     "website": "https://www.optima-packaging.com", "central_email": "info@optima-packaging.com", "central_phone": "+49 791 9495-0",
     "email_domain": "optima-packaging.com", "pattern_hint": "{first}.{last}@optima-packaging.com",
     "produkte_kurz": "Pharma/Cosmetics/Consumer-Linien, Serialization, Digital Twin, AI",
     "zielkunden_kurz": "Pfizer, Roche, P&G, Henkel, L'Oréal",
     "branche_kontext": "Verpackung (V): GMP + PPWR; Engineering Change Orders.",
     "pain_kundensprache": "GMP-Varianten vor Freigabe in belastbare Kosten?",
     "contacts": [
         {"type": "CEO/GF", "name": "Dr. Stefan König", "role": "CEO (seit Apr 2024)",
          "domain": "optima-packaging.com", "primary_email": "stefan.koenig@optima-packaging.com", "extras": [],
          "source_url": "https://www.optima-packaging.com/en/newsroom/press/press-releases/dr.-stefan-koenig-is-the-new-ceo-of-the-optima-group",
          "source_type": "presse",
          "website_fakt": "CEO-Wechsel 2024; interpack 2026 AI/Digital Twin; Fiber-Packaging; ~3.400 MA.",
          "hook_fakt": "GMP-Varianten und PPWR rutschen bei Pharma-/Consumer-Linien oft früher in die Angebotsphase — institutionalisiert oder in Köpfen?",
          "hook_quelle": "https://www.optima-packaging.com/en/newsroom/press/",
          "betreff": "interpack-Nachlese Optima — GMP-Varianten früher in RFQs?"},
         {"type": "Engineering", "name": "Tobias Meinikheim", "role": "Director of Engineering",
          "domain": "optima-packaging.com", "extras": ["tobias.meinikheim@optima-packaging.com"],
          "source_url": "https://www.linkedin.com/in/tobias-meinikheim-63b66217b", "source_type": "linkedin",
          "website_fakt": "Director Engineering seit 2024; Digital Twin/AI auf interpack 2026.",
          "hook_fakt": "Bei GMP-Linien: Trifft Nacharbeit nach Freigabe zu — oder stockt die Varianten-Durchlaufzeit in der Konstruktion?",
          "hook_quelle": "https://www.optima-packaging.com",
          "betreff": "GMP-Engineering — Freigabe oder Nacharbeit als Engpass?"},
     ]},
    # === 9 Schubert ===
    {"nr": 9, "segment": "Packaging", "company": "Gerhard Schubert GmbH", "location": "DE, Crailsheim",
     "website": "https://www.schubert.group", "central_email": "info@schubert.group", "central_phone": "+49 7951 400-0",
     "email_domain": "gerhard-schubert.de", "pattern_hint": "{f}.{last}@gerhard-schubert.de",
     "produkte_kurz": "TLM Pick-and-Place, F4-Roboter, Schubert-Pharma, Vision/AI",
     "zielkunden_kurz": "Food, Pharma, Cosmetics, Medical Devices",
     "branche_kontext": "Verpackung (V): Modulare Linien, hohe Variantenkomplexität.",
     "pain_kundensprache": "Modulare Linien = Variantenvielfalt — RFQ-Tempo vs. Engineering-Kapazität?",
     "contacts": [
         {"type": "CEO/GF", "name": "Ralf Schubert", "role": "Geschäftsführer (Technologie-Division)",
          "domain": "gerhard-schubert.de", "primary_email": "r.schubert@gerhard-schubert.de", "extras": [],
          "source_url": "https://www.schubert.group/en/imprint/", "source_type": "impressum",
          "website_fakt": "Schubert-Pharma Expansion; modulares TLM-System; AI Vision.",
          "hook_fakt": "Modulare TLM-Linien erzeugen hohe Variantenvielfalt — limitiert RFQ-Tempo Kapazität oder Marge pro Auftrag?",
          "hook_quelle": "https://www.schubert.group",
          "betreff": "TLM-Varianten — RFQ-Tempo oder Marge limitiert?"},
         {"type": "Engineering", "name": "Marcus Schindler", "role": "Bereichsleiter Konstruktion",
          "domain": "gerhard-schubert.de", "primary_email": "m.schindler@gerhard-schubert.de",
          "extras": ["t.neumann@gerhard-schubert.de"],
          "source_url": "https://www.schubert.group/en/group/locations/", "source_type": "firmenwebsite",
          "website_fakt": "Pick-and-Place TLM; variantenreiche Food-/Pharma-Linien.",
          "hook_fakt": "Bei modularen Linien: Werden Material- und Formatvarianten vor der Freigabe in derselben Kostensicht bewertet?",
          "hook_quelle": "https://www.schubert.group",
          "betreff": "Modulare Linien — Kostensicht vor Freigabe?"},
     ]},
    # === 10 Bielomatik ===
    {"nr": 10, "segment": "Packaging", "company": "Bielomatik GmbH (Persico Group)", "location": "DE, Römerstein",
     "website": "https://www.bielomatik.com", "central_email": "info@bielomatik.de", "central_phone": "+49 7382 9427-0",
     "email_domain": "bielomatik.de", "pattern_hint": "{first}.{last}@bielomatik.de",
     "produkte_kurz": "Plastic Joining Systems (Paper-Sparte → BW Papersystems seit 2015)",
     "zielkunden_kurz": "Automotive, Industrial Joining",
     "branche_kontext": "Hinweis: Paper-Converting an BW Papersystems verkauft; aktuell Plastic Welding.",
     "pain_kundensprache": "Nach Restrukturierung: variantenreiche Angebote — Kalkulation institutionalisiert?",
     "contacts": [
         {"type": "CEO/GF", "name": "Marco Imperatore", "role": "Geschäftsführer",
          "domain": "bielomatik.de", "extras": ["marco.imperatore@bielomatik.de", "marco.imperatore@bielomatik.com"],
          "source_url": "https://www.bielomatik.com/en/imprint.html", "source_type": "impressum",
          "website_fakt": "Paper-Sparte 2015 an BW Papersystems; heute Plastic Joining unter Persico, Römerstein.",
          "hook_fakt": "Nach dem Verkauf der Paper-Sparte an BW Papersystems: Wie kalkulieren Sie heute variantenreiche Joining-Systeme — institutionalisiert oder projektweise?",
          "hook_quelle": "https://www.barrywehmiller.com/news/company-news/release/barry-wehmiller-acquires-bielomatik-paper-processing-group",
          "betreff": "Nach Paper-Sparten-Verkauf — wie läuft Kalkulation heute?"},
         {"type": "Engineering", "name": "Robin Brunzendorf", "role": "Key Account Manager / techn. Ansprechpartner Europa",
          "domain": "persico.com", "extras": ["robin.brunzendorf@persico.com"],
          "source_url": "https://www.bielomatik.com/en/contact.html", "source_type": "firmenwebsite",
          "website_fakt": "Persico-Gruppe; Plastic Welding Service: service-plasticwelding@bielomatik.de.",
          "hook_fakt": "Bei kundenspezifischen Joining-Lösungen: Varianten vor Freigabe in belastbare Kosten — oder Nacharbeit?",
          "hook_quelle": "https://www.bielomatik.com/en/contact.html",
          "betreff": "Joining-Varianten — Freigabe oder Nacharbeit?",
          "remarks_extra": "Eng-Kontakt über Persico-Gruppe (verifizierte Domain persico.com)"},
     ]},
    # === 11 EREMA ===
    {"nr": 11, "segment": "Packaging", "company": "EREMA Engineering Recycling Maschinen", "location": "AT, Ansfelden",
     "website": "https://www.erema.com", "central_email": "erema@erema.at", "central_phone": "+43 732 3190-0",
     "email_domain": "erema.at", "pattern_hint": "{first}.{last}@erema.at / @erema.com",
     "produkte_kurz": "INTAREMA/VACUREMA PCR-Recycling, Film-Linien, PURE LOOP",
     "zielkunden_kurz": "Packaging Producer, Recycler, FMCG mit PCR-Zielen",
     "branche_kontext": "Verpackung/Recycling (V): PPWR, PCR-Anteil als Kundenspec.",
     "pain_kundensprache": "PCR-Anteil als Kundenspec — Maschinenvariante im Design kalkulierbar?",
     "contacts": [
         {"type": "CEO/GF", "name": "Manfred Hackl", "role": "CEO / Geschäftsführer",
          "domain": "erema.at", "extras": ["manfred.hackl@erema.at", "manfred.hackl@erema.com", "m.hackl@erema.at"],
          "source_url": "https://www.erema.com/de/impressum/", "source_type": "impressum",
          "website_fakt": "Marktführer Plastic Recycling; Way2K/K 2025 Edvanced Recycling; PPWR-Treiber.",
          "hook_fakt": "PCR-Anteil als Kundenspec treibt Maschinenvarianten — rutschen diese früher in Ihre Angebotsphase als früher?",
          "hook_quelle": "https://edvanced.erema.com/en/way2k-2025-interview-with-manfred-hackl/",
          "betreff": "PPWR & PCR-Specs — früher in EREMA-Angeboten?"},
         {"type": "Engineering", "name": "Markus Huber-Lindinger", "role": "Geschäftsführer / Technik",
          "domain": "erema.at", "extras": ["markus.huber-lindinger@erema.at", "markus.huber-lindinger@erema.com"],
          "source_url": "https://www.erema.com/de/impressum/", "source_type": "impressum",
          "website_fakt": "Drei GF (Hackl, Huber-Lindinger, Wolfsgruber); variantenreiche Recycling-Linien.",
          "hook_fakt": "Bei PCR-Recycling-Linien: Werden Materialmix-Varianten vor der Freigabe in derselben Kostensicht bewertet?",
          "hook_quelle": "https://www.erema.com/de/impressum/",
          "betreff": "Materialmix-Varianten — vor Freigabe kalkuliert?"},
     ]},
    # === 12 Bobst ===
    {"nr": 12, "segment": "Packaging", "company": "Bobst Group SA", "location": "CH, Mex",
     "website": "https://www.bobst.com", "central_email": "info@bobst.com", "central_phone": "",
     "email_domain": "bobst.com", "pattern_hint": "{first}.{last}@bobst.com",
     "produkte_kurz": "Folding Carton, Corrugated, Flexible Packaging Machinery",
     "zielkunden_kurz": "Converter, Brand Owner, Global Packaging",
     "branche_kontext": "Verpackung (V) Konzern: Konfigurationsvielfalt; Presse-Routing.",
     "pain_kundensprache": "Konfigurationsvielfalt — Bottleneck Angebot vs. Engineering?",
     "contacts": [
         {"type": "CEO/GF", "name": "Jean-Pascal Bobst", "role": "CEO Bobst Group",
          "domain": "bobst.com", "extras": ["jean-pascal.bobst@bobst.com", "j.bobst@bobst.com", "gudrun.alex@bobst.com"],
          "source_url": "https://bobst.prezly.com/", "source_type": "presse",
          "website_fakt": "Inventor Award 2025; Digital/IoT; Sustainability; CHF 1.6B Umsatz.",
          "hook_fakt": "Bei Konfigurationsvielfalt über Divisionen: Bottleneck Angebot oder Engineering-Kapazität?",
          "hook_quelle": "https://bobst.prezly.com/bobst-inventor-award-2025-recognizes-exceptional-employee-drive-and-innovation",
          "betreff": "Bobst 2026 — Angebots- oder Engineering-Engpass?",
          "routing_primary": "gudrun.alex@bobst.com"},
         {"type": "Engineering", "name": "Leonard Badet", "role": "Chief Technology Officer",
          "domain": "bobst.com", "extras": ["leonard.badet@bobst.com", "l.badet@bobst.com", "gudrun.alex@bobst.com"],
          "source_url": "https://www.linkedin.com/in/leonard-badet/en", "source_type": "linkedin",
          "website_fakt": "CTO seit 2023; >500 R&D-Mitarbeiter global; Digital Transformation.",
          "hook_fakt": "Konfigurationsvarianten in Folding Carton/Flexible: Nacharbeit nach Freigabe oder Varianten-Durchlaufzeit?",
          "hook_quelle": "https://bobst.prezly.com/",
          "betreff": "R&D-Konfigurationen — Freigabe oder Durchlaufzeit?"},
     ]},
    # === 13 Senning ===
    {"nr": 13, "segment": "Packaging", "company": "Senning GmbH", "location": "DE, Bremen",
     "website": "https://www.senning.de", "central_email": "info@senning.de", "central_phone": "+49 421 694620",
     "email_domain": "senning.de", "pattern_hint": "{first}.{last}@senning.de",
     "produkte_kurz": "Tissue/Paper Verpackungsmaschinen, ETO",
     "zielkunden_kurz": "Confectionery, Tobacco, Tissue, Food",
     "branche_kontext": "Verpackung (V): Optima-Gruppe; ~50 MA; 85% Export.",
     "pain_kundensprache": "Kleines ETO — Kalkulation am GF/Ingenieur gebunden?",
     "contacts": [
         {"type": "CEO/GF", "name": "Annette Bengs", "role": "Geschäftsführerin (3. Generation)",
          "domain": "senning.de", "extras": ["annette.bengs@senning.de", "a.bengs@senning.de"],
          "source_url": "https://www.senning.de/impressum", "source_type": "impressum",
          "website_fakt": "Optima-Gruppe seit 2020; Familienunternehmen; 85% Export Tissue/Paper.",
          "hook_fakt": "Als familiengeführtes ETO mit ~50 MA: Bindet Kalkulation am GF/Ingenieur — oder ist das institutionalisiert?",
          "hook_quelle": "https://www.optima-packaging.com/de/newsroom/presse/pressemitteilungen/senning-wird-teil-der-optima-unternehmensgruppe",
          "betreff": "Senning als ETO — Kalkulation in Köpfen oder System?"},
         {"type": "Engineering", "name": "Oliver Rebstock", "role": "Geschäftsführer / Operations",
          "domain": "senning.de", "extras": ["oliver.rebstock@senning.de", "o.rebstock@senning.de"],
          "source_url": "https://www.senning.de/impressum", "source_type": "impressum",
          "website_fakt": "Tissue-/Paper-Verpackungsmaschinen; ETO pro Kunde; Bremen.",
          "hook_fakt": "Bei Tissue-Linien: Varianten vor Freigabe in belastbare Kosten — oder Nacharbeit?",
          "hook_quelle": "https://www.senning.de",
          "betreff": "Tissue-Varianten — vor Freigabe kalkuliert?"},
     ]},
    # === 14 Fill ===
    {"nr": 14, "segment": "Packaging", "company": "Fill GmbH", "location": "AT, Gurten",
     "website": "https://www.fill.co.at", "central_email": "info@fill.co.at", "central_phone": "+43 7757 7010",
     "email_domain": "fill.co.at", "pattern_hint": "{first}.{last}@fill.co.at (lehre@ verifiziert)",
     "produkte_kurz": "Sondermaschinenbau Metall/Kunststoff/Holz — kein klassischer Verpackungs-OEM",
     "zielkunden_kurz": "Automotive, Aviation, Sport, Bau",
     "branche_kontext": "Sondermaschinenbau (E); Hexagon-Humanoid-Partnerschaft 2026.",
     "pain_kundensprache": "ETO-Projektgeschäft — wo stocken RFQ-Schleifen?",
     "contacts": [
         {"type": "CEO/GF", "name": "Andreas Fill", "role": "CEO",
          "domain": "fill.co.at", "extras": ["andreas.fill@fill.co.at", "a.fill@fill.co.at"],
          "source_url": "https://www.fill.co.at", "source_type": "firmenwebsite",
          "website_fakt": "EY-Award Innovationsführer; Hexagon-Humanoid-Partnerschaft 2026; 11,6% F&E-Quote.",
          "hook_fakt": "Mit Humanoid-Partnerschaft und hoher F&E-Quote: Wo stocken kundenspezifische RFQ-Schleifen — Variantenkalkulation oder Freigabe?",
          "hook_quelle": "https://www.fill.co.at/en/news-and-events/media/innovation-winner-from-the-innviertel-region",
          "betreff": "Humanoid-Projekte Fill — RFQ-Durchlaufzeit oder Marge?"},
         {"type": "Engineering", "name": "Alois Wiesinger", "role": "CTO",
          "domain": "fill.co.at", "extras": ["alois.wiesinger@fill.co.at", "a.wiesinger@fill.co.at"],
          "source_url": "https://www.fill.co.at", "source_type": "firmenwebsite",
          "website_fakt": "Sondermaschinen-ETO; hohe Variantenkomplexität; >1.000 MA.",
          "hook_fakt": "Bei Sondermaschinen-ETO: Trifft Nacharbeit nach Freigabe zu — oder Varianten-Durchlaufzeit?",
          "hook_quelle": "https://www.fill.co.at",
          "betreff": "Sondermaschinen-Varianten — Freigabe oder Nacharbeit?"},
     ]},
]
# fmt: on


def last_name(full: str) -> str:
    _, la = parse_name(full)
    return la


def salutation(name: str) -> str:
    if not name:
        return "Guten Tag,"
    female = {"annette", "cristina", "sabine", "petra"}
    first = parse_name(name)[0].lower()
    ln = last_name(name)
    if first in female:
        return f"Guten Tag Frau {ln},"
    return f"Guten Tag Herr {ln},"


def company_short(company: str) -> str:
    c = company.split("(")[0].strip()
    for cut in (" GmbH", " KG", " SA", " Group", " & Co."):
        if cut in c:
            c = c.split(cut)[0].strip()
    return c.split()[0] if c else company


def build_ceo_body(row: dict) -> str:
    short = company_short(row["company"])
    seg = row["segment"]
    hook = row["hook_fakt"]
    intro = (
        "ich schreibe Ihnen, weil wir bei LoopForgeLab mit wenigen ETO-Herstellern "
        "im Verpackungsmaschinenbau den Austausch zu Kalkulation und Varianten im Design "
        "suchen — ohne IT-Großprojekt."
        if seg == "Packaging" and "Fill" not in row["company"]
        else "ich schreibe Ihnen, weil wir bei LoopForgeLab mit wenigen ETO-Herstellern "
        "in der Intralogistik und im Sondermaschinenbau den Austausch zu Kalkulation "
        "und Varianten im Design suchen — ohne IT-Großprojekt."
    )
    return f"""{salutation(row["name"])}

{intro}

{hook}

Aus Gesprächen mit anderen Geschäftsführern im Sondermaschinenbau höre ich oft eine Spannung — ich würde gern prüfen, ob das bei Ihnen überhaupt zutrifft:

Entweder limitiert **Angebotskapazität** das Wachstum (zu viel Engineering in der Kalkulation), oder die **Marge** pro gewonnenem Auftrag (Puffer, Nachkalkulation, zu späte Kostensicht).

**Was ist bei {short} 2026 der schwerere Engpass — oder liegt Ihr Fokus ganz woanders?**

Eine kurze Einordnung reicht mir — ich möchte nicht ins Leere schreiben.

Beste Grüße
Olaf Pick
LoopForgeLab GbR
olaf@loopforgelab.com"""


def build_eng_body(row: dict) -> str:
    short = company_short(row["company"])
    hook = row["hook_fakt"]
    return f"""{salutation(row["name"])}

ich gehe gezielt auf Engineering-Leitungen im Sondermaschinenbau zu — Hintergrund: wie Varianten und Kosten **vor der Freigabe** zusammenlaufen.

{hook}

In anderen ETO-Betrieben höre ich häufig: Die teuersten Überraschungen kommen nicht in der Montage selbst, sondern **vor der Freigabe** — wenn Varianten und Materialwechsel noch nicht mit derselben Kostensicht bewertet werden.

**Trifft „Nacharbeit / Änderungsaufträge nach Freigabe“ bei {short} eher zu — oder ist der Engpass eher die Durchlaufzeit der Varianten in der Konstruktion?**

Ich bin an Ihrer Einordnung interessiert, nicht an einem Produkt-Pitch.

Beste Grüße
Olaf Pick
LoopForgeLab GbR
olaf@loopforgelab.com"""


def routing_body(row: dict) -> str:
    name, role = row["name"], row["role"]
    core = build_ceo_body(row) if row["contact_type"] == "CEO/GF" else build_eng_body(row)
    inner = core.split("\n", 2)[2] if core.startswith("Guten Tag") else core
    for cut in ("Beste Grüße", "Best regards"):
        if cut in inner:
            inner = inner.split(cut)[0].strip()
    target = f"{name} ({role})" if name else role
    return f"""Guten Tag,

könnten Sie diese Nachricht bitte an {target} weiterleiten?

---

{inner}

---

Vielen Dank für die Weiterleitung.

Beste Grüße
Olaf Pick
LoopForgeLab GbR
olaf@loopforgelab.com"""


def mx_hosts(domain: str) -> list[str]:
    try:
        import dns.resolver  # type: ignore
        answers = dns.resolver.resolve(domain, "MX")
        pairs = sorted((r.preference, str(r.exchange).rstrip(".")) for r in answers)
        return [h for _, h in pairs]
    except Exception:
        pass
    try:
        socket.getaddrinfo(domain, 25, proto=socket.IPPROTO_TCP)
        return [domain]
    except socket.gaierror:
        return []


def smtp_check(email: str) -> dict:
    email = email.strip().lower()
    if email in SMTP_CACHE:
        return SMTP_CACHE[email]
    if not email or "@" not in email:
        r = {"email": email, "status": "skip", "detail": "ungültig"}
        SMTP_CACHE[email] = r
        return r
    domain = email.split("@", 1)[1]
    hosts = mx_hosts(domain)
    if not hosts:
        r = {"email": email, "status": "fail", "detail": "kein MX"}
        SMTP_CACHE[email] = r
        return r
    last_err = ""
    for host in hosts[:3]:
        try:
            with smtplib.SMTP(host, 25, timeout=TIMEOUT) as smtp:
                smtp.ehlo_or_helo_if_needed()
                smtp.mail(FROM_ADDR)
                code, msg = smtp.rcpt(email)
                text = msg.decode(errors="replace") if isinstance(msg, bytes) else str(msg)
                if code in (250, 251):
                    r = {"email": email, "status": "ok", "detail": f"{host}: {code}"}
                    SMTP_CACHE[email] = r
                    return r
                if code in (550, 551, 553, 554):
                    detail = f"{host}: {code} {text[:120]}"
                    if any(x in text.lower() for x in ("user unknown", "recipient not found", "does not exist", "no such user")):
                        r = {"email": email, "status": "fail", "detail": detail, "reason": "user_unknown"}
                    elif any(x in text.lower() for x in ("spamhaus", "blocked", "client host", "barracuda")):
                        r = {"email": email, "status": "blocked", "detail": detail, "reason": "ip_block"}
                    else:
                        r = {"email": email, "status": "fail", "detail": detail, "reason": "reject"}
                    SMTP_CACHE[email] = r
                    return r
                last_err = f"{host}: {code} {text[:100]}"
        except Exception as e:
            last_err = f"{host}: {type(e).__name__}"
            time.sleep(0.2)
    r = {"email": email, "status": "unknown", "detail": last_err or "timeout"}
    SMTP_CACHE[email] = r
    return r


def pick_best_email(candidates: list[str], routing: str | None = None) -> tuple[str, str, list[dict], str]:
    """Returns email, status label, all results, variants_tested string."""
    tested = []
    to_test = [c for c in candidates if c and not c.startswith("info@") and not c.startswith("office@")][:MAX_VARIANTS_SMTP]
    for em in to_test:
        if em not in [t["email"] for t in tested]:
            tested.append(smtp_check(em))
            time.sleep(0.1)
    ok = [t for t in tested if t["status"] == "ok"]
    blocked = [t for t in tested if t["status"] == "blocked"]
    user_fail = [t for t in tested if t.get("reason") == "user_unknown"]
    if ok:
        best = ok[0]["email"]
        label = f"smtp OK ({best})"
        status = "verifiziert (SMTP)"
        return best, status, tested, "; ".join(f"{t['email']}:{t['status']}" for t in tested)
    if routing:
        return routing, "routing — Presse/Zentrale", tested, "; ".join(f"{t['email']}:{t['status']}" for t in tested)
    if blocked and not user_fail:
        best = candidates[0] if candidates else ""
        return best, "abgeleitet (SMTP blockiert — IP)", tested, "; ".join(f"{t['email']}:{t['status']}" for t in tested)
    if tested:
        best = candidates[0]
        t0 = tested[0]
        if t0.get("reason") == "user_unknown":
            # try next candidate
            for t in tested[1:]:
                if t["status"] == "ok":
                    return t["email"], "verifiziert (SMTP)", tested, ""
            return best, "abgeleitet (SMTP user unknown)", tested, ""
        return best, "abgeleitet (mittel)", tested, "; ".join(f"{t['email']}:{t['status']}" for t in tested)
    return candidates[0] if candidates else "", "nicht getestet", tested, ""


def smtp_label_from_results(email: str, tested: list[dict], is_routing: bool) -> str:
    if is_routing:
        oks = [t for t in tested if t["status"] == "ok"]
        if oks:
            return f"routing — Direkt OK: {oks[0]['email']}"
        blocked = [t for t in tested if t["status"] == "blocked"]
        if blocked and len(blocked) == len([t for t in tested if t["status"] in ("blocked", "fail")]):
            return "routing — SMTP IP-blockiert; Adresse plausibel"
        return "routing — manuell prüfen"
    for t in tested:
        if t["email"] == email.lower() and t["status"] == "ok":
            return "smtp OK"
    for t in tested:
        if t["email"] == email.lower():
            if t["status"] == "blocked":
                return "smtp IP-blockiert — Adresse plausibel, nicht verifizierbar"
            if t.get("reason") == "user_unknown":
                return f"smtp FAIL user unknown — {email}"
            if t["status"] == "fail":
                return f"smtp FAIL — {t['detail'][:70]}"
            return f"smtp UNBEKANNT — {t['detail'][:70]}"
    return "nicht getestet"


def build_rows() -> tuple[list[dict], list[dict]]:
    rows, smtp_log = [], []
    for co in CONTACTS:
        for ct in co["contacts"]:
            fi, la = parse_name(ct["name"])
            domain = ct["domain"]
            candidates = build_candidates(fi, la, domain, ct.get("extras"))
            if ct.get("primary_email"):
                pe = ct["primary_email"].strip().lower()
                candidates = [pe] + [c for c in candidates if c != pe]
            routing = ct.get("routing_primary")
            is_routing = bool(routing)
            email, email_status, tested, variants_str = pick_best_email(candidates, routing)
            smtp_lbl = smtp_label_from_results(email, tested, is_routing)
            hook_lens = "GF_kapazitaet_marge" if ct["type"] == "CEO/GF" else "Eng_varianten_freigabe"
            row = {
                "nr": co["nr"], "segment": co["segment"], "company": co["company"],
                "location": co["location"], "website": co["website"],
                "contact_type": ct["type"], "name": ct["name"], "role": ct["role"],
                "email": email, "email_status": email_status,
                "email_pattern": co["pattern_hint"],
                "smtp_check": smtp_lbl,
                "email_variants_tested": variants_str,
                "source_url": ct["source_url"], "source_type": ct["source_type"],
                "alt_email": co["central_email"],
                "central_email": co["central_email"], "central_phone": co.get("central_phone", ""),
                "linkedin_url": ct.get("linkedin_url", ""),
                "produkte_kurz": co["produkte_kurz"], "zielkunden_kurz": co["zielkunden_kurz"],
                "branche_kontext": co["branche_kontext"],
                "website_fakt": ct["website_fakt"], "hook_fakt": ct["hook_fakt"],
                "hook_quelle": ct["hook_quelle"], "hook_lens": hook_lens,
                "pain_kundensprache": co["pain_kundensprache"],
                "email_betreff": ct["betreff"],
                "status": "routing_erforderlich" if is_routing else "recherche_fertig",
                "remarks": ct.get("remarks_extra", ""),
            }
            if ct["type"] == "CEO/GF":
                row["email_body"] = build_ceo_body(row)
            else:
                row["email_body"] = build_eng_body(row)
            row["email_body_routing"] = routing_body(row) if is_routing else ""
            rows.append(row)
            smtp_log.append({"nr": co["nr"], "name": ct["name"], "email": email, "smtp_check": smtp_lbl, "tested": tested})
            print(f"  [{co['nr']}] {ct['name'][:28]:28} {email[:35]:35} {smtp_lbl}", flush=True)
    return rows, smtp_log


def style_sheet(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    seg_pack = PatternFill("solid", fgColor="E2EFDA")
    seg_intra = PatternFill("solid", fgColor="DDEBF7")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    widths = {"company": 34, "hook_fakt": 44, "email_body": 42, "email_body_routing": 42,
              "email_variants_tested": 36, "remarks": 28, "email_betreff": 32}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 14)
    for r in range(2, ws.max_row + 1):
        seg = ws.cell(r, headers.index("segment") + 1).value
        fill = seg_pack if seg == "Packaging" else seg_intra
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(r, c).fill = fill


def write_sheet(wb, rows: list[dict]) -> None:
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    for c, h in enumerate(HEADERS, 1):
        ws.cell(1, c, h)
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(HEADERS, 1):
            ws.cell(r, c, row.get(h, ""))
    style_sheet(ws, HEADERS)
    last_col = get_column_letter(len(HEADERS))
    tab = Table(displayName="KontakteEmail", ref=f"A1:{last_col}{len(rows)+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)


def write_standalone_xlsx(rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kontakte_Email"
    for c, h in enumerate(HEADERS, 1):
        ws.cell(1, c, h)
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(HEADERS, 1):
            ws.cell(r, c, row.get(h, ""))
    style_sheet(ws, HEADERS)
    last_col = get_column_letter(len(HEADERS))
    tab = Table(displayName="KontakteEmailExport", ref=f"A1:{last_col}{len(rows)+1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUT)


def write_csv(rows: list[dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with CSV_OUT.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in HEADERS})


def main() -> None:
    print("Baue Kontakte + SMTP-Varianten …")
    rows, smtp_log = build_rows()
    print(f"\n{len(rows)} Zeilen — schreibe Dateien …")
    for path in (MASTER, KONTAKTE_COPY):
        if path == KONTAKTE_COPY or path.exists():
            try:
                wb = openpyxl.load_workbook(path if path.exists() else MASTER)
                write_sheet(wb, rows)
                wb.save(path)
                print(f"  OK: {path}")
            except PermissionError:
                print(f"  GESPERRT: {path}")
    write_standalone_xlsx(rows)
    print(f"  OK: {XLSX_OUT}")
    write_csv(rows)
    print(f"  OK: {CSV_OUT}")
    SMTP_LOG.write_text(json.dumps(smtp_log, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    ok_count = sum(1 for r in rows if r["smtp_check"] == "smtp OK" or "Direkt OK" in r["smtp_check"])
    print(f"\nFertig: {ok_count}/{len(rows)} mit smtp OK")


if __name__ == "__main__":
    main()
