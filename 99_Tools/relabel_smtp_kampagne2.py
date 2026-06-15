#!/usr/bin/env python3
"""Relabel smtp_check from saved JSON (Spamhaus vs real fail)."""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "99_Tools"))
from apply_research_kampagne2 import is_ip_block, smtp_label, parse_alt_emails  # noqa: E402

XLSX = ROOT / "260615_Kampagne_2.xlsx"
JSON = ROOT / "04_Outreach/Kampagnen/2026-06-15_Kampagne_2/_smtp_results_kampagne2.json"
SHEET = "Kontakte_Email"


def main():
    log = json.loads(JSON.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    status_col = {h: i + 1 for i, h in enumerate(headers) if h == "status"}
    for entry in log:
        nr = entry["nr"]
        name = entry["name"]
        for r in range(2, ws.max_row + 1):
            if int(ws.cell(r, col["nr"]).value) != nr:
                continue
            if ws.cell(r, col["name"]).value != name:
                continue
            primary = (entry.get("email") or "").lower()
            tested = entry.get("tested", [])
            status = ws.cell(r, col["status"]).value or ""
            is_routing = "routing" in str(status) or (primary or "").startswith("info@")
            label = smtp_label(primary, tested, is_routing)
            ws.cell(r, col["smtp_check"], label)
            print(f"  {name}: {label}")
    wb.save(XLSX)
    print(f"Updated {XLSX}")


if __name__ == "__main__":
    main()
